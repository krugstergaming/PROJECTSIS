from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment
from django.utils.timezone import now
from django.contrib.auth.hashers import make_password
from decimal import Decimal, InvalidOperation
from django.db import IntegrityError
from django.db.models import Count, F
from django.db import transaction
from datetime import datetime, date
from bs4 import BeautifulSoup
from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse, HttpResponseRedirect, JsonResponse
from django.contrib import messages
from django.urls import reverse
from django.views.decorators.csrf import csrf_exempt
import json
import cloudinary.uploader
import openpyxl
import requests

from student_management_app.models import CustomUser, Staffs, StudentPromotionHistory, Curriculums, GradeLevel, Enrollment, Attachment, BalancePayment, Subjects, Section, AssignSection, Load, Schedule, Students, SessionYearModel, FeedBackStudent, FeedBackStaffs, LeaveReportStudent, LeaveReportStaff, Attendance, AttendanceReport, GradingConfiguration, ParentGuardian, PreviousSchool, EmergencyContact, Classroom
from .forms import EditStudentForm, AddScheduleForm, EditScheduleForm


def admin_home(request):
    all_student_count = Students.objects.count()
    subject_count = Subjects.objects.count()
    gradelevel_count = GradeLevel.objects.count()
    staff_count = Staffs.objects.count()
    load_count = Load.objects.count()
    curriculum_count = Curriculums.objects.count()

    # Serialize GradeLevel data
    grade_levels = GradeLevel.objects.values('GradeLevel_name')
    grade_levels_data = json.dumps(list(grade_levels))  # Serialize data for JSON compatibility

    # Updated query to get subjects grouped by GradeLevel
    subjects_per_gradelevel = Subjects.objects.select_related('GradeLevel_id') \
        .values('GradeLevel_id__GradeLevel_name') \
        .annotate(subject_count=Count('id'))

    # Prepare data to be passed to Chart.js
    gradelevels_datas = [
        {
            "grade_level": item['GradeLevel_id__GradeLevel_name'],  # The grade level name
            "subject_count": item['subject_count']  # The count of subjects per grade level
        }
        for item in subjects_per_gradelevel
    ]

    # Aggregate the number of students per GradeLevel
    students_per_gradelevel = Students.objects.values('GradeLevel_id__GradeLevel_name').annotate(student_count=Count('id'))
    
    # Prepare the data to pass to the template
    student_gradelevel_data = [
        {
            "grade_level": item['GradeLevel_id__GradeLevel_name'],
            "student_count": item['student_count']
        }
        for item in students_per_gradelevel
    ]

    context = {
        "all_student_count": all_student_count,
        "subject_count": subject_count,
        "gradelevel_count": gradelevel_count,
        "staff_count": staff_count,
        "load_count": load_count,
        "curriculum_count": curriculum_count,
        "curriculum_count": curriculum_count,
        "grade_levels_data": grade_levels_data,  # Pass JSON-serialized data to context
        "gradelevels_datas": json.dumps(gradelevels_datas),  # Pass JSON-serialized data to context
        "student_gradelevel_data": json.dumps(student_gradelevel_data),  # Pass JSON-serialized data to context
    }

    return render(request, "hod_template/home_content.html", context)




def add_staff(request):
    return render(request, "hod_template/Add_Template/add_staff_template.html")

def add_staff_save(request):
    if request.method != "POST":
        messages.error(request, "Invalid Method ")
        return redirect('add_staff')
    else:
        try:
            first_name = request.POST.get('first_name')
            last_name = request.POST.get('last_name')
            username = request.POST.get('username')
            email = request.POST.get('email')
            max_laod = request.POST.get('max_load')
            middle_name = request.POST.get('middle_name')
            dob = request.POST.get('dob')
            age = request.POST.get('age')
            pob = request.POST.get('pob')
            sex = request.POST.get('sex')
            civil_status = request.POST.get('civil_status')
            height = request.POST.get('height')
            weight = request.POST.get('weight')
            blood_type = request.POST.get('blood_type')
            gsis_id = request.POST.get('gsis_id')
            pagibig_id = request.POST.get('pagibig_id')
            philhealth_id = request.POST.get('philhealth_id')
            sss_id = request.POST.get('sss_id')
            tin_id = request.POST.get('tin_id')
            citizenship = request.POST.get('citizenship')
            dual_country = request.POST.get('dual_country')
            permanent_address = request.POST.get('permanent_address')
            telephone_no = request.POST.get('telephone_no')
            cellphone_no = request.POST.get('cellphone_no')

            # Generate a predefined password based on staff details
            current_year = now().year
            predefined_password = f"{first_name.lower()}.{last_name.lower()}.{current_year}"

            # Create Staff User
            user = CustomUser.objects.create_user(
                username=username,
                password=make_password(predefined_password),  # Use hashed password
                email=email,
                first_name=first_name,
                last_name=last_name,
                user_type=2
            )
            user.staffs.max_laod = max_laod
            user.staffs.middle_name = middle_name
            user.staffs.dob = dob
            user.staffs.age = age
            user.staffs.pob = pob
            user.staffs.sex = sex
            user.staffs.civil_status = civil_status
            user.staffs.height = height
            user.staffs.weight = weight
            user.staffs.blood_type = blood_type
            user.staffs.gsis_id = gsis_id
            user.staffs.pagibig_id = pagibig_id
            user.staffs.philhealth_id = philhealth_id
            user.staffs.sss_id = sss_id
            user.staffs.tin_id = tin_id
            user.staffs.citizenship = citizenship
            user.staffs.dual_country = dual_country
            user.staffs.permanent_address = permanent_address
            user.staffs.telephone_no = telephone_no
            user.staffs.cellphone_no = cellphone_no
            user.save()
            
            messages.success(request, "Faculty Added Successfully! Password: " + predefined_password)
            return redirect('add_staff')
        except Exception as e:
            messages.error(request, f"Failed to Add Staff: {str(e)}")
            return redirect('add_staff')

def toggle_grading_state(request):
    grading_config, created = GradingConfiguration.objects.get_or_create(id=1)
    # Toggle the grading state
    grading_config.is_grading_active = not grading_config.is_grading_active
    grading_config.save()
    
    # Redirect back to the manage_staff view
    return redirect('manage_staff')

def manage_staff(request):
    staffs = Staffs.objects.all()
    grading_config, created = GradingConfiguration.objects.get_or_create(id=1)
    context = {
        "staffs": staffs,
        "grading_config": grading_config,  # Pass the grading configuration to the template
    }
    return render(request, "hod_template/Manage_Template/manage_staff_template.html", context)


def edit_staff(request, staff_id):
    staff = Staffs.objects.get(admin=staff_id)

    context = {
        "staff": staff,
        "id": staff_id
    }
    return render(request, "hod_template/Edit_Template/edit_staff_template.html", context)


def edit_staff_save(request):
    if request.method != "POST":
        return HttpResponse("<h2>Method Not Allowed</h2>")
    else:
        staff_id = request.POST.get('staff_id')
        username = request.POST.get('username')
        email = request.POST.get('email')
        first_name = request.POST.get('first_name')
        last_name = request.POST.get('last_name')
        address = request.POST.get('address')

        try:
            # INSERTING into Customuser Model
            user = CustomUser.objects.get(id=staff_id)
            user.first_name = first_name
            user.last_name = last_name
            user.email = email
            user.username = username
            user.save()
            
            # INSERTING into Staff Model
            staff_model = Staffs.objects.get(admin=staff_id)
            staff_model.address = address
            staff_model.save()

            messages.success(request, "Staff Updated Successfully.")
            return redirect('/edit_staff/'+staff_id)

        except:
            messages.error(request, "Failed to Update Staff.")
            return redirect('/edit_staff/'+staff_id)


def deactivate_staff(request, staff_id):
    try:
        # Get the CustomUser object related to the staff
        user = CustomUser.objects.get(id=staff_id)

        # Set the user as inactive
        user.is_active = False
        user.save()

        messages.success(request, "Staff account deactivated successfully.")
        return redirect('manage_staff')
    except CustomUser.DoesNotExist:
        messages.error(request, "Staff not found.")
        return redirect('manage_staff')
    except Exception as e:
        messages.error(request, f"Failed to deactivate staff: {str(e)}")
        return redirect('manage_staff')



def add_curriculum(request):
    return render(request, "hod_template/Add_Template/add_curriculum_template.html")


def add_curriculum_save(request):
    if request.method != "POST":
        messages.error(request, "Invalid Method!")
        return redirect('add_curriculum')
    else:
        curriculum_name = request.POST.get('curriculum_name')
        curriculum_description = request.POST.get('curriculum_description')
        curriculum_status = request.POST.get('curriculum_status')
        try:
            curriculum = Curriculums(
                curriculum_name=curriculum_name,
                curriculum_description=curriculum_description,
                curriculum_status=curriculum_status
                )
            curriculum.save()
            messages.success(request, "Curriculum Added Successfully!")
            return redirect('add_curriculum')
        except:
            messages.error(request, "Failed to Add Curriculum!")
            return redirect('add_curriculum')
        

def manage_curriculum(request):
    curriculums = Curriculums.objects.all()
    curriculums = Curriculums.objects.filter(is_archived=False)
    context = {
        "curriculums": curriculums
    }
    return render(request, 'hod_template/Manage_Template/manage_curriculum_template.html', context)

def edit_curriculum(request, curriculum_id):
    curriculum = Curriculums.objects.get(id=curriculum_id)
    context = {
        "curriculum": curriculum,
        "id": curriculum_id
    }
    return render(request, 'hod_template/Edit_Template/edit_curriculum_template.html', context)

def edit_curriculum_save(request):
    if request.method != "POST":
        HttpResponse("Invalid Method")
    else:
        curriculum_id = request.POST.get('curriculum_id')
        curriculum_name = request.POST.get('curriculum_name')
        curriculum_description = request.POST.get('curriculum_description')
        curriculum_status = request.POST.get('curriculum_status')
        try:
            curriculum = Curriculums.objects.get(id=curriculum_id)
            curriculum.curriculum_name = curriculum_name
            curriculum.curriculum_description = curriculum_description
            curriculum.curriculum_status = curriculum_status
            curriculum.save()
            messages.success(request, "Curriculum Updated Successfully.")
            return redirect('/manage_curriculum/')
        except:
            messages.error(request, "Failed to Update GradeLevel.")
            return redirect('/manage_curriculum/')
def archived_curriculums(request):
    # Filter curriculums with status 'Archived'
    curriculums = Curriculums.objects.filter(is_archived=True)
    context = {
        "curriculums": curriculums
    }
    return render(request, 'hod_template/Archive_Template/archived_curriculum_template.html', context)

def archive_curriculum(request, curriculum_id):
    curriculum = get_object_or_404(Curriculums, id=curriculum_id)
    curriculum.is_archived = True
    curriculum.curriculum_status = "Inactive"
    curriculum.save()
    messages.success(request, "Curriculum archived successfully!")
    return redirect('/manage_curriculum/')  
def unarchive_curriculum(request, curriculum_id):
    curriculum = get_object_or_404(Curriculums, id=curriculum_id)
    curriculum.is_archived = False
    curriculum.curriculum_status = "Active"
    curriculum.save()
    messages.success(request, "Curriculum unarchived successfully!")
    return redirect('/manage_curriculum/')  


def add_gradelevel(request):
    curriculums = Curriculums.objects.all()
    context = {
        "curriculums": curriculums
    }
    return render(request, "hod_template/Add_Template/add_gradelevel_template.html", context)

def add_gradelevel_save(request):
    if request.method != "POST":
        messages.error(request, "Invalid Method!")
        return redirect('add_gradelevel')
    else:

        GradeLevel_name = request.POST.get('GradeLevel_name')

        curriculum_id = request.POST.get('curriculum_id')
        curriculums = Curriculums.objects.get(id=curriculum_id)

        try:
            gradelevel = GradeLevel(
                GradeLevel_name=GradeLevel_name,
                curriculum_id = curriculums
                )
            gradelevel.save()
            messages.success(request, "GradeLevel Added Successfully!")
            return redirect('add_gradelevel')
        except:
            messages.error(request, "Failed to Add GradeLevel!")
            return redirect('add_gradelevel')
        
def export_gradelevels_to_excel(request):
    # Create an in-memory workbook
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = 'Grade Levels'
    
    # Define the headers
    headers = ['GradeLevel Name', 'Curriculum', 'Date Created', 'Date Updated']
    
    # Define a fill color for the header (e.g., light blue)
    header_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
    
    # Add headers to the first row and make them bold with a background color and center alignment
    for col_num, header in enumerate(headers, 1):
        cell = sheet.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)  # Make the header bold
        cell.fill = header_fill  # Apply background color
        cell.alignment = Alignment(horizontal="center", vertical="center")  # Center-align the text

    # Query all grade levels
    gradelevels = GradeLevel.objects.all()

    # Add data rows starting from row 2 with center alignment
    for row_num, gradelevel in enumerate(gradelevels, 2):
        sheet.cell(row=row_num, column=1).value = gradelevel.GradeLevel_name
        sheet.cell(row=row_num, column=2).value = gradelevel.curriculum_id.curriculum_name  # Assuming curriculum has 'curriculum_name'
        sheet.cell(row=row_num, column=3).value = gradelevel.created_at.strftime('%Y-%m-%d %H:%M:%S')
        sheet.cell(row=row_num, column=4).value = gradelevel.updated_at.strftime('%Y-%m-%d %H:%M:%S')

        # Apply center alignment to all cells in the row
        for col_num in range(1, 5):
            cell = sheet.cell(row=row_num, column=col_num)
            cell.alignment = Alignment(horizontal="center", vertical="center")

    # Adjust column widths to fit the content
    for col in sheet.columns:
        max_length = 0
        column = col[0].column_letter  # Get the column letter (A, B, C, etc.)
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = max_length + 2  # Add some padding to the width
        sheet.column_dimensions[column].width = adjusted_width

    # Prepare the response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=GradeLevels.xlsx'

    # Save the workbook to the response
    workbook.save(response)

    return response

def export_staffs_to_excel(request):
    # Create an in-memory workbook
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = 'Staffs'
    
    # Define the headers
    headers = ['Last Name', 'First Name', 'Middle Name', 'Email', 'Age', 'Date of Birth', 'Place of Birth', 
               'Sex', 'Civil Rights', 'Height', 'Weight', 'Blood Type', 'GSIG ID', 'PagIbig ID', 
               'PhilHealth ID', 'SSS ID', 'TIN ID', 'Citizenship', 'Address', 'Telephone #',
               'Cellphone #', 'Created At', 'Updated At', 'Is Active']
    
    # Define a fill color for the header (e.g., light blue)
    header_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
    
    # Add headers to the first row and make them bold with a background color and center alignment
    for col_num, header in enumerate(headers, 1):
        cell = sheet.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)  # Make the header bold
        cell.fill = header_fill  # Apply background color
        cell.alignment = Alignment(horizontal="center", vertical="center")  # Center-align the text

    # Query all grade levels
    staffs = Staffs.objects.all()

    # Add data rows starting from row 2 with center alignment
    for row_num, staff in enumerate(staffs, 2):
        sheet.cell(row=row_num, column=1).value = staff.admin.last_name
        sheet.cell(row=row_num, column=2).value = staff.admin.first_name
        sheet.cell(row=row_num, column=3).value = staff.middle_name
        sheet.cell(row=row_num, column=4).value = staff.admin.email
        sheet.cell(row=row_num, column=5).value = staff.age
        sheet.cell(row=row_num, column=6).value = staff.dob
        sheet.cell(row=row_num, column=7).value = staff.pob
        sheet.cell(row=row_num, column=8).value = staff.sex
        sheet.cell(row=row_num, column=9).value = staff.civil_status
        sheet.cell(row=row_num, column=10).value = staff.height
        sheet.cell(row=row_num, column=11).value = staff.weight
        sheet.cell(row=row_num, column=12).value = staff.blood_type
        sheet.cell(row=row_num, column=13).value = staff.gsis_id
        sheet.cell(row=row_num, column=14).value = staff.pagibig_id
        sheet.cell(row=row_num, column=15).value = staff.philhealth_id
        sheet.cell(row=row_num, column=16).value = staff.sss_id
        sheet.cell(row=row_num, column=17).value = staff.tin_id
        sheet.cell(row=row_num, column=18).value = staff.citizenship
        sheet.cell(row=row_num, column=19).value = staff.permanent_address
        sheet.cell(row=row_num, column=20).value = staff.telephone_no
        sheet.cell(row=row_num, column=21).value = staff.cellphone_no
        sheet.cell(row=row_num, column=22).value = staff.created_at.strftime('%Y-%m-%d %H:%M:%S')
        sheet.cell(row=row_num, column=23).value = staff.updated_at.strftime('%Y-%m-%d %H:%M:%S')
        sheet.cell(row=row_num, column=24).value = staff.admin.is_active

        # Apply center alignment to all cells in the row
        for col_num in range(1, 25):
            cell = sheet.cell(row=row_num, column=col_num)
            cell.alignment = Alignment(horizontal="center", vertical="center")

    # Adjust column widths to fit the content
    for col in sheet.columns:
        max_length = 0
        column = col[0].column_letter  # Get the column letter (A, B, C, etc.)
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = max_length + 2  # Add some padding to the width
        sheet.column_dimensions[column].width = adjusted_width

    # Prepare the response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=StaffRecord.xlsx'

    # Save the workbook to the response
    workbook.save(response)

    return response

def export_students_to_excel(request):
    # Create an in-memory workbook
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = 'Students'
    
    # Define the headers
    headers = ['Last Name', 'First Name', 'Middle Name', 'Email', 'Grade Level', 'Age', 'Date of Birth', 'Place of Birth', 
               'Sex', 'Nationality', 'Religion', 'Rank In Family', 'Telephone #',
               'Mobile #', 'Start Semester', 'End Semester', 'Created At', 'Updated At', 'Is Active']
    
    # Define a fill color for the header (e.g., light blue)
    header_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
    
    # Add headers to the first row and make them bold with a background color and center alignment
    for col_num, header in enumerate(headers, 1):
        cell = sheet.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)  # Make the header bold
        cell.fill = header_fill  # Apply background color
        cell.alignment = Alignment(horizontal="center", vertical="center")  # Center-align the text

    # Query all grade levels
    students = Students.objects.all()

    # Add data rows starting from row 2 with center alignment
    for row_num, student in enumerate(students, 2):
        sheet.cell(row=row_num, column=1).value = student.admin.last_name
        sheet.cell(row=row_num, column=2).value = student.admin.first_name
        sheet.cell(row=row_num, column=3).value = student.middle_name
        sheet.cell(row=row_num, column=4).value = student.admin.email
        sheet.cell(row=row_num, column=5).value = student.GradeLevel_id.GradeLevel_name
        sheet.cell(row=row_num, column=6).value = student.age
        sheet.cell(row=row_num, column=7).value = student.dob
        sheet.cell(row=row_num, column=8).value = student.sex
        sheet.cell(row=row_num, column=9).value = student.pob
        sheet.cell(row=row_num, column=10).value = student.nationality
        sheet.cell(row=row_num, column=11).value = student.religion
        sheet.cell(row=row_num, column=12).value = student.rank_in_family
        sheet.cell(row=row_num, column=13).value = student.telephone_nos
        sheet.cell(row=row_num, column=14).value = student.mobile_phone_nos
        sheet.cell(row=row_num, column=15).value = student.session_year_id.session_start_year
        sheet.cell(row=row_num, column=16).value = student.session_year_id.session_end_year
        sheet.cell(row=row_num, column=17).value = student.created_at.strftime('%Y-%m-%d %H:%M:%S')
        sheet.cell(row=row_num, column=18).value = student.updated_at.strftime('%Y-%m-%d %H:%M:%S')
        sheet.cell(row=row_num, column=19).value = student.admin.is_active

        # Apply center alignment to all cells in the row
        for col_num in range(1, 20):
            cell = sheet.cell(row=row_num, column=col_num)
            cell.alignment = Alignment(horizontal="center", vertical="center")

    # Adjust column widths to fit the content
    for col in sheet.columns:
        max_length = 0
        column = col[0].column_letter  # Get the column letter (A, B, C, etc.)
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = max_length + 2  # Add some padding to the width
        sheet.column_dimensions[column].width = adjusted_width

    # Prepare the response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=StudentRecord.xlsx'

    # Save the workbook to the response
    workbook.save(response)

    return response


def manage_gradelevel(request):
    gradelevels = GradeLevel.objects.all()
    context = {
        "gradelevels": gradelevels
    }
    return render(request, 'hod_template/Manage_Template/manage_gradelevel_template.html', context)


def edit_gradelevel(request, GradeLevel_id):
    gradelevel = GradeLevel.objects.get(id=GradeLevel_id)
    context = {
        "gradelevel": gradelevel,
        "id": GradeLevel_id
    }
    return render(request, 'hod_template/Edit_Template/edit_gradelevel_template.html', context)


def edit_gradelevel_save(request):
    if request.method != "POST":
        HttpResponse("Invalid Method")
    else:
        GradeLevel_id = request.POST.get('GradeLevel_id')
        GradeLevel_name = request.POST.get('gradelevel')

        try:
            gradelevel = GradeLevel.objects.get(id=GradeLevel_id)
            gradelevel.GradeLevel_name = GradeLevel_name
            gradelevel.save()

            messages.success(request, "GradeLevel Updated Successfully.")
            return redirect('/edit_gradelevel/'+GradeLevel_id)

        except:
            messages.error(request, "Failed to Update GradeLevel.")
            return redirect('/edit_gradelevel/'+GradeLevel_id)

def delete_gradelevel(request, GradeLevel_id):
    gradelevel = GradeLevel.objects.get(id=GradeLevel_id)
    try:
        gradelevel.delete()
        messages.success(request, "GradeLevel Deleted Successfully.")
        return redirect('manage_gradelevel')
    except:
        messages.error(request, "Failed to Delete GradeLevel.")
        return redirect('manage_gradelevel')


def manage_session(request):
    session_years = SessionYearModel.objects.filter(is_archived=False)
    context = {
        "session_years": session_years
    }
    return render(request, "hod_template/Manage_Template/manage_session_template.html", context)

def add_session(request):
    # Get the latest session year if it exists
    latest_session = SessionYearModel.objects.order_by('-session_end_year').first()
    
    if latest_session:
        # Use the end year of the last session as the start year of the new session
        last_end_year = latest_session.session_end_year.year
        # Set the new session years based on the last session
        session_start_year = f"{last_end_year}-01-01"  # Start of the new session year
        session_end_year = f"{last_end_year + 1}-12-31"  # End of the new session year
    else:
        # If no session exists, set default values, e.g., for the current year
        current_year = datetime.now().year
        session_start_year = f"{current_year}-01-01"
        session_end_year = f"{current_year + 1}-12-31"

    print(f"Session Start Year: {session_start_year}, Session End Year: {session_end_year}")  # Debug print

    return render(request, "hod_template/Add_Template/add_session_template.html", {
        'session_start_year': session_start_year,
        'session_end_year': session_end_year
    })

def add_session_save(request):
    if request.method != "POST":
        messages.error(request, "Invalid Method")
        return redirect('add_session')
    else:
        session_start_year = request.POST.get('session_start_year')
        session_end_year = request.POST.get('session_end_year')
        session_limit = request.POST.get('session_limit')
        session_status = "Active"

        try:
            sessionyear = SessionYearModel(
                session_start_year = session_start_year, 
                session_end_year = session_end_year,
                session_limit = session_limit,
                session_status = session_status
                )
            sessionyear.save()
            messages.success(request, "Session Year added Successfully!")
            return redirect("add_session")
        except:
            messages.error(request, "Failed to Add Session Year")
            return redirect("add_session")

def edit_session(request, session_id):
    session_year = SessionYearModel.objects.get(id=session_id)
    context = {
        "session_year": session_year
    }
    return render(request, "hod_template/Edit_Template/edit_session_template.html", context)

def edit_session_save(request):
    if request.method != "POST":
        messages.error(request, "Invalid Method!")
        return redirect('manage_session')
    else:
        session_id = request.POST.get('session_id')
        session_start_year = request.POST.get('session_start_year')
        session_end_year = request.POST.get('session_end_year')
        session_limit = request.POST.get('session_limit')
        session_status = request.POST.get('session_status')

        try:
            session_year = SessionYearModel.objects.get(id=session_id)
            session_year.session_start_year = session_start_year
            session_year.session_end_year = session_end_year
            session_year.session_limit = session_limit
            session_year.session_status = session_status
            session_year.save()

            messages.success(request, "Session Year Updated Successfully.")
            return redirect('manage_session')
        except Exception as e:
            messages.error(request, f"Failed to Update Session Year: {e}")
            return redirect('manage_session')

def archived_sessions(request):
    # Filter session years with status 'Archived'
    session_years = SessionYearModel.objects.filter(is_archived=True)
    context = {
        "session_years": session_years
    }
    return render(request, 'hod_template/Archive_Template/archived_session_template.html', context)

def archive_session(request, session_id):
    session_year = get_object_or_404(SessionYearModel, id=session_id)
    session_year.is_archived = True
    session_year.session_status = "Inactive"
    session_year.save()
    messages.success(request, "Session archived successfully!")
    return redirect('manage_session')

def unarchive_session(request, session_id):
    session_year = get_object_or_404(SessionYearModel, id=session_id)
    session_year.is_archived = False
    session_year.session_status = "Active"
    session_year.save()
    messages.success(request, "Session unarchived successfully!")
    return redirect('manage_session')

def delete_session(request, session_id):
    session = SessionYearModel.objects.get(id=session_id)
    try:
        session.delete()
        messages.success(request, "Session Deleted Successfully.")
        return redirect('manage_session')
    except:
        messages.error(request, "Failed to Delete Session.")
        return redirect('manage_session')
    
def manage_classroom(request):
    classrooms = Classroom.objects.all()

    # Define context with headers, data, and statistics
    context = {
        "classrooms": classrooms
    }

    return render(request, "hod_template/Manage_Template/manage_classroom_template.html", context)

def add_classroom(request):
    if request.method != "POST":
        messages.error(request, "Method Not Allowed!")
        return redirect('manage_classroom')
    try:
        # Fetch and parse the table data from the external website
        r = requests.get('https://logisticsinteg.onrender.com/Depart_Inv.html')
        soup = BeautifulSoup(r.text, 'html.parser')
        table = soup.find('table')
        rows = table.find_all('tr')
        
        # Extract data rows, skipping headers
        data = []
        for row in rows[1:]:  # Skips the first row (headers)
            cols = row.find_all('td')
            data.append([col.text.strip() for col in cols])

        # Fetch existing classroom names to avoid duplicates
        existing_names = set(Classroom.objects.values_list('classroom_name', flat=True))
        new_classrooms = []

        # Identify and prepare unique Classroom objects
        for row_data in data:
            if row_data:
                classroom_name = row_data[0]
                if classroom_name not in existing_names:
                    new_classrooms.append(Classroom(classroom_name=classroom_name))
                    existing_names.add(classroom_name)

        # Bulk create new unique records within a transaction
        with transaction.atomic():
            Classroom.objects.bulk_create(new_classrooms)

        messages.success(request, "Classroom Fetched Successfully!")
        return redirect('manage_classroom')
    
    except Exception as e:
        messages.error(request, f"Failed to Fetch Classroom! Error: {str(e)}")
        return redirect('manage_classroom')

def add_student(request):
    gradelevels = GradeLevel.objects.all()
    sessions = SessionYearModel.objects.all()
    year = datetime.now().year

    # Debug print to check current year
    print(f"Current Year: {year}")

    # Fetch the last student number starting with the current year
    last_student = Students.objects.filter(student_number__startswith=f"{year}-").order_by('-student_number').first()

    if last_student:
        # Debug print to check the last student number fetched
        print(f"Last Student Number: {last_student.student_number}")
        
        # Extract the last four digits after the dash, convert to an integer, and increment by 1
        last_number = int(last_student.student_number.split('-')[1])
        new_number = f"{year}-{str(last_number + 1).zfill(4)}"
    else:
        # Start with 0001 if no student exists for the current year
        new_number = f"{year}-0001"

    # Debug print to check the new student number
    print(f"New Student Number: {new_number}")

    context = {
        "gradelevels": gradelevels,
        "sessions": sessions,
        "student_number": new_number
    }
    return render(request, 'hod_template/Add_Template/add_student2_template.html', context)

def add_student_save(request):
    if request.method != "POST":
        messages.error(request, "Method Not Allowed!")
        return redirect('add_student')

    try:
        # Extract data from request.POST
        first_name = request.POST.get('first_name')
        last_name = request.POST.get('last_name')
        username = request.POST.get('username')
        middlename = request.POST.get('middle_name')

        suffix = request.POST.get("suffix")
        nickname = request.POST.get("nickname")

        email = request.POST.get('email')
        address = request.POST.get('address')
        session_year_id = request.POST.get('session_year_id')
        GradeLevel_id = request.POST.get('GradeLevel_id')
        sex = request.POST.get('sex')
        age = request.POST.get('age')
        dob = request.POST.get('dob')
        pob = request.POST.get('pob')
        nationality = request.POST.get('nationality')
        religion = request.POST.get('religion')
        rank_in_family = request.POST.get('rank_in_family')
        telephone_nos = request.POST.get('telephone_nos')
        mobile_phone_nos = request.POST.get('mobile_phone_nos')
        is_covid_vaccinated = request.POST.get('is_covid_vaccinated')
        date_of_vaccination = request.POST.get('date_of_vaccination')

        # Handle file upload for profile picture
        profile_pic_url = None
        if 'profile_pic' in request.FILES:
            profile_pic = request.FILES['profile_pic']
            upload_result = cloudinary.uploader.upload(profile_pic, folder="media/")
            profile_pic_url = upload_result.get('url')

        # Generate a predefined password based on the student's details
        current_year = now().year
        predefined_password = f"{first_name.lower()}.{last_name.lower()}.{current_year}"

        # Generate student_number
        last_student = Students.objects.filter(student_number__startswith=str(current_year)).order_by('-id').first()
        if last_student:
            last_number = int(last_student.student_number.split('-')[1])
            new_number = last_number + 1
        else:
            new_number = 1
        student_number = f"{current_year}-{new_number:04d}"

        # Create Student User
        user = CustomUser.objects.create_user(
            username=username, 
            password=make_password(predefined_password),  # Use hashed password
            email=email,
            first_name=first_name, 
            last_name=last_name, 
            user_type=3
        )
        user.students.student_number = student_number  # Set student_number
        user.students.address = address

        # Assign GradeLevel and Session Year
        gradelevel_obj = GradeLevel.objects.get(id=GradeLevel_id)
        session_year_obj = SessionYearModel.objects.get(id=session_year_id)
        user.students.GradeLevel_id = gradelevel_obj
        user.students.session_year_id = session_year_obj

        # Additional Student Details
        user.students.middle_name = middlename
        user.students.suffix = suffix
        user.students.nickname = nickname
        user.students.sex = sex
        user.students.profile_pic = profile_pic_url
        user.students.age = age
        user.students.dob = dob
        user.students.pob = pob
        user.students.nationality = nationality
        user.students.religion = religion
        user.students.rank_in_family = rank_in_family
        user.students.telephone_nos = telephone_nos
        user.students.mobile_phone_nos = mobile_phone_nos
        user.students.is_covid_vaccinated = is_covid_vaccinated
        user.students.date_of_vaccination = date_of_vaccination
        user.save()

        # Save Parent/Guardian Information
        ParentGuardian.objects.create(
            students_id=user.students,
            father_name=request.POST.get('father_name'),
            father_occupation=request.POST.get('father_occupation'),
            mother_name=request.POST.get('mother_name'),
            mother_occupation=request.POST.get('mother_occupation'),
            guardian_name=request.POST.get('guardian_name'),
            guardian_occupation=request.POST.get('guardian_occupation')
        )

        # Save Previous School Information
        PreviousSchool.objects.create(
            students_id=user.students,
            previous_school_name=request.POST.get('previous_school_name'),
            previous_school_address=request.POST.get('previous_school_address'),
            previous_grade_level=request.POST.get('previous_grade_level'),
            previous_school_year_attended=request.POST.get('previous_school_year_attended'),
            previous_teacher_name=request.POST.get('previous_teacher_name')
        )

        # Save Emergency Contact Information
        EmergencyContact.objects.create(
            students_id=user.students,
            emergency_contact_name=request.POST.get('emergency_contact_name'),
            emergency_contact_relationship=request.POST.get('emergency_contact_relationship'),
            emergency_contact_address=request.POST.get('emergency_contact_address'),
            emergency_contact_phone=request.POST.get('emergency_contact_phone'),
            emergency_enrolling_teacher=request.POST.get('emergency_enrolling_teacher'),
            emergency_date=request.POST.get('emergency_date'),
            emergency_referred_by=request.POST.get('emergency_referred_by')
        )

        messages.success(request, "Student Added Successfully! Password: " + predefined_password)
        return redirect('add_student')

    except Exception as e:
        messages.error(request, f"Failed to Add Student! Error: {str(e)}")
        return redirect('add_student')

def add_enrollment(request):
    # Get all students who are NOT in the Enrollment table
    enrolled_students_ids = Enrollment.objects.values_list('student_id', flat=True)
    students = Students.objects.exclude(id__in=enrolled_students_ids)
    
    context = {
        "students": students
    }
    return render(request, 'hod_template/Add_Template/add_enrollment_template.html', context)

def add_enrollment_save(request):
    if request.method != "POST":
        messages.error(request, "Invalid Method!")
        return redirect('add_enrollment')

    # Retrieve student instance
    student_id = request.POST.get('student_id')
    
    try:
        student_instance = Students.objects.get(id=student_id)
    except Students.DoesNotExist:
        messages.error(request, "Student not found.")
        return redirect('add_enrollment')

    # Check if the student already has an enrollment record
    if Enrollment.objects.filter(student_id=student_instance).exists():
        messages.error(request, "Student is already enrolled.")
        return redirect('add_enrollment')

    # Retrieve and convert values from the form
    def safe_decimal(value, field_name):
        """Helper function to safely convert to Decimal with error logging."""
        try:
            return Decimal(value.strip() or '0.00')
        except InvalidOperation:
            messages.error(request, f"Invalid input for {field_name}. Please check your values.")
            raise

    try:
        # Convert and log values
        registration_fee = safe_decimal(request.POST.get('registration_fee', '0.00'), 'registration fee')
        misc_fee = safe_decimal(request.POST.get('misc_fee', '0.00'), 'misc fee')
        tuition_fee = safe_decimal(request.POST.get('tuition_fee', '0.00'), 'tuition fee')
        total_fee = safe_decimal(request.POST.get('total_fee', '0.00'), 'total fee')
        downpayment = safe_decimal(request.POST.get('downpayment', '0.00'), 'downpayment')
        discount = safe_decimal(request.POST.get('discount', '0.00'), 'discount')
        discount_amount = safe_decimal(request.POST.get('discount_amount', '0.00'), 'discount amount')
        balance = safe_decimal(request.POST.get('balance', '0.00'), 'balance')
        installment_payment = safe_decimal(request.POST.get('installment_payment', '0.00'), 'installment payment')
        payment_amount = safe_decimal(request.POST.get('payment_amount', '0.00'), 'payment amount')

        installment_option = request.POST.get('installment_option', 'Monthly')
        assessed_by = request.POST.get('assessed_by', '')
        assessed_date = request.POST.get('assessed_date', None)  # Modify as needed for date format
        payment_received_by = request.POST.get('payment_received_by', '')
        payment_date = request.POST.get('payment_date', None)  # Modify as needed for date format
        enrollment_status = request.POST.get('enrollment_status', 'Pending')
        remarks = request.POST.get('remarks', '')

        # Create and save the enrollment instance
        enrollment = Enrollment(
            student_id=student_instance,
            registration_fee=registration_fee,
            misc_fee=misc_fee,
            tuition_fee=tuition_fee,
            total_fee=total_fee,
            downpayment=downpayment,
            discount=discount,
            discount_amount=discount_amount,
            balance=balance,
            installment_payment=installment_payment,
            installment_option=installment_option,
            assessed_by=assessed_by,
            assessed_date=assessed_date,
            payment_received_by=payment_received_by,
            payment_amount=payment_amount,
            payment_date=payment_date,
            enrollment_status=enrollment_status,
            remarks=remarks,
        )
        
        enrollment.save()

        # Update student status to "Enrolled"
        student_instance.student_status = "Enrolled"
        student_instance.save()

        # Handle file uploads from the request
        id_picture_file = request.FILES.get('id_picture_file') if request.POST.get('include_id_picture') else None
        psa_file = request.FILES.get('birth_certificate_file') if request.POST.get('include_birth_certificate') else None
        form_138_file = request.FILES.get('form_138_file') if request.POST.get('include_form_138') else None

        # Create an Attachment instance
        attachment = Attachment(
            enrollment=enrollment,
            id_picture_file=id_picture_file,
            psa_file=psa_file,
            form_138_file=form_138_file,
            attachment_remarks=request.POST.get('attachment_remarks', '')
        )

        attachment.save()
        messages.success(request, "Enrollment Added Successfully!")
    except Exception as e:
        messages.error(request, f"Failed to Add Enrollment! Error: {str(e)}")
    return redirect('add_enrollment')

def manage_enrollment(request):
    students = Students.objects.all()
    enrollments = Enrollment.objects.all()
    attachments = Attachment.objects.all()
    balancepayments = BalancePayment.objects.all()
    context = {
        "students": students,
        "enrollments": enrollments,
        "attachments": attachments,
        "balancepayments": balancepayments,
    }
    return render(request, 'hod_template/Manage_Template/manage_enrollment_template.html', context)

def edit_enrollment(request, enrollment_id):
    # Retrieve the enrollment instance
    enrollment = get_object_or_404(Enrollment, id=enrollment_id)
    
    # Prepare context with existing data from the enrollment instance
    context = {
        "enrollment_id": enrollment.id,
        "student_id": enrollment.student_id.id,
        "student_name": f"{enrollment.student_id.admin.first_name} {enrollment.student_id.admin.last_name}",
        "registration_fee": enrollment.registration_fee,
        "misc_fee": enrollment.misc_fee,
        "tuition_fee": enrollment.tuition_fee,
        "total_fee": enrollment.total_fee,
        "downpayment": enrollment.downpayment,
        "discount": enrollment.discount,
        "discount_amount": enrollment.discount_amount,
        "balance": enrollment.balance,
        "installment_payment": enrollment.installment_payment,
        "installment_option": enrollment.installment_option,
        "assessed_by": enrollment.assessed_by,
        "assessed_date": enrollment.assessed_date,
        "payment_received_by": enrollment.payment_received_by,
        "payment_amount": enrollment.payment_amount,
        "payment_date": enrollment.payment_date,
        "enrollment_status": enrollment.enrollment_status,
        "remarks": enrollment.remarks,
    }

    # Check for attachments and add them to the context if they exist
    attachment = getattr(enrollment, 'attachment', None)
    if attachment is not None:
        context.update({
            "id_picture_file": attachment.id_picture_file or '',  # Provide default if None
            "birth_certificate_file": attachment.psa_file or '',  # Provide default if None
            "form_138_file": attachment.form_138_file or '',  # Provide default if None
            "attachment_remarks": attachment.attachment_remarks or '',  # Provide default if None
        })
    else:
        # If no attachment exists, you might want to set default values for the context
        context.update({
            "id_picture_file": '',
            "birth_certificate_file": '',
            "form_138_file": '',
            "attachment_remarks": '',
        })

    # Render the edit template with populated data
    return render(request, "hod_template/Edit_Template/edit_enrollment_template.html", context)

def edit_enrollment_save(request, enrollment_id):
    if request.method != "POST":
        messages.error(request, "Invalid Method!")
        return redirect('edit_enrollment', enrollment_id=enrollment_id)

    # Retrieve the existing enrollment instance
    enrollment = get_object_or_404(Enrollment, id=enrollment_id)
    student_instance = enrollment.student_id

    # Retrieve and convert values from the form
    def safe_decimal(value, field_name):
        """Helper function to safely convert to Decimal with error logging."""
        try:
            return Decimal(value.strip() or '0.00')
        except InvalidOperation:
            messages.error(request, f"Invalid input for {field_name}. Please check your values.")
            raise

    try:
        # Convert and log values
        enrollment.registration_fee = safe_decimal(request.POST.get('registration_fee', '0.00'), 'registration fee')
        enrollment.misc_fee = safe_decimal(request.POST.get('misc_fee', '0.00'), 'misc fee')
        enrollment.tuition_fee = safe_decimal(request.POST.get('tuition_fee', '0.00'), 'tuition fee')
        enrollment.total_fee = safe_decimal(request.POST.get('total_fee', '0.00'), 'total fee')
        enrollment.downpayment = safe_decimal(request.POST.get('downpayment', '0.00'), 'downpayment')
        enrollment.discount = safe_decimal(request.POST.get('discount', '0.00'), 'discount')
        enrollment.discount_amount = safe_decimal(request.POST.get('discount_amount', '0.00'), 'discount amount')
        enrollment.balance = safe_decimal(request.POST.get('balance', '0.00'), 'balance')
        enrollment.installment_payment = safe_decimal(request.POST.get('installment_payment', '0.00'), 'installment payment')
        enrollment.payment_amount = safe_decimal(request.POST.get('payment_amount', '0.00'), 'payment amount')

        enrollment.installment_option = request.POST.get('installment_option', 'Monthly')
        enrollment.assessed_by = request.POST.get('assessed_by', '')
        enrollment.assessed_date = request.POST.get('assessed_date', None)
        enrollment.payment_received_by = request.POST.get('payment_received_by', '')
        enrollment.payment_date = request.POST.get('payment_date', None)
        enrollment.enrollment_status = request.POST.get('enrollment_status', 'Pending')
        enrollment.remarks = request.POST.get('remarks', '')

        # Save the updated enrollment instance
        enrollment.save()

        # Handle file uploads from the request
        if request.POST.get('include_id_picture'):
            enrollment.attachment.id_picture_file = request.FILES.get('id_picture_file')
        if request.POST.get('include_birth_certificate'):
            enrollment.attachment.psa_file = request.FILES.get('birth_certificate_file')
        if request.POST.get('include_form_138'):
            enrollment.attachment.form_138_file = request.FILES.get('form_138_file')

        enrollment.attachment.attachment_remarks = request.POST.get('attachment_remarks', '')
        enrollment.attachment.save()

        messages.success(request, "Enrollment Updated Successfully!")
    except Exception as e:
        messages.error(request, f"Failed to Update Enrollment! Error: {str(e)}")

    return redirect('edit_enrollment', enrollment_id=enrollment_id)

def update_balance(request, enrollment_id=None):
    # Fetch enrollment based on enrollment_id
    enrollment = get_object_or_404(Enrollment, id=enrollment_id)

    if request.method == 'POST':
        # Get form data from POST request
        payment_balance_amount = request.POST.get('payment_balance_amount')
        payment_balance_date = request.POST.get('payment_balance_date')
        payment_balance_remarks = request.POST.get('payment_balance_remarks')

        try:
            # Remove commas and convert to Decimal
            payment_balance_amount = Decimal(payment_balance_amount.replace(',', ''))

            # Store previous balance
            past_balance = enrollment.balance

            # Update the balance (ensure both are Decimal for arithmetic)
            enrollment.balance -= payment_balance_amount
            enrollment.save()

            # Create a new payment record with past balance
            BalancePayment.objects.create(
                enrollment=enrollment,
                payment_balance_amount=payment_balance_amount,
                payment_balance_date=payment_balance_date,
                payment_balance_remarks=payment_balance_remarks,
                past_balance=past_balance
            )

            # Success message
            messages.success(request, f'Balance updated successfully. Current balance: {enrollment.balance:.2f}, Past balance: {past_balance:.2f}')
            return redirect('manage_enrollment')

        except IntegrityError:
            messages.error(request, 'There was an error updating the balance. Please try again.')
        except Exception as e:
            messages.error(request, f'An unexpected error occurred: {str(e)}')

    return render(request, 'hod_template/Manage_Template/manage_enrollment_template.html', {
        'enrollment': enrollment,
        'enrollments': Enrollment.objects.all(),
    })  
 
# def promote_students(request):
#     if request.method == "POST":
#         current_grade_id = request.POST.get('current_grade')
#         next_grade_id = request.POST.get('next_grade')
        
#         try:
#             current_grade = GradeLevel.objects.get(id=current_grade_id)
#             next_grade = GradeLevel.objects.get(id=next_grade_id)
            
#             # Get all students in the current grade
#             students_to_promote = Students.objects.filter(GradeLevel_id=current_grade)
            
#             # Create a promotion history for each student
#             for student in students_to_promote:
#                 StudentPromotionHistory.objects.create(
#                     student=student,
#                     previous_grade=current_grade,
#                     new_grade=next_grade,
#                     promotion_date=timezone.now()
#                 )

#             # Promote students to the next grade
#             students_to_promote.update(GradeLevel_id=next_grade)
            
#             messages.success(request, f"All students have been promoted from {current_grade.name} to {next_grade.name}.")
#         except Exception as e:
#             messages.error(request, f"Failed to promote students. Error: {e}")
        
#     return redirect('students_list')

# def view_promotion_history(request, student_id):
#     promotion_history = StudentPromotionHistory.objects.filter(student_id=student_id)
#     return render(request, 'view_promotion_history.html', {'promotion_history': promotion_history})


def manage_student(request):
    students = Students.objects.all()
    context = {
        "students": students
    }
    return render(request, 'hod_template/Manage_Template/manage_student_template.html', context)


def edit_student(request, student_id):
    # Adding Student ID into Session Variable
    request.session['student_id'] = student_id

    student = Students.objects.get(admin=student_id)
    form = EditStudentForm()
    # Filling the form with Data from Database
    form.fields['email'].initial = student.admin.email
    form.fields['username'].initial = student.admin.username
    form.fields['first_name'].initial = student.admin.first_name
    form.fields['last_name'].initial = student.admin.last_name
    form.fields['address'].initial = student.address
    form.fields['GradeLevel_id'].initial = student.GradeLevel_id.id
    form.fields['sex'].initial = student.sex
    form.fields['session_year_id'].initial = student.session_year_id.id

    context = {
        "id": student_id,
        "username": student.admin.username,
        "form": form
    }
    return render(request, "hod_template/Edit_Template/edit_student_template.html", context)


def edit_student_save(request):
    if request.method != "POST":
        return HttpResponse("Invalid Method!")
    else:
        student_id = request.session.get('student_id')
        if student_id == None:
            return redirect('/manage_student')

        form = EditStudentForm(request.POST, request.FILES)
        if form.is_valid():
            email = form.cleaned_data['email']
            username = form.cleaned_data['username']
            first_name = form.cleaned_data['first_name']
            last_name = form.cleaned_data['last_name']
            address = form.cleaned_data['address']
            GradeLevel_id = form.cleaned_data['GradeLevel_id']
            gender = form.cleaned_data['gender']
            session_year_id = form.cleaned_data['session_year_id']

            # Getting Profile Pic first
            # First Check whether the file is selected or not
            # Upload only if file is selected
            profile_pic_url = None
            if 'profile_pic' in request.FILES:
                profile_pic = request.FILES['profile_pic']
                upload_result = cloudinary.uploader.upload(profile_pic, folder="media/")
                profile_pic_url = upload_result.get('url')

            try:
                # First Update into Custom User Model
                user = CustomUser.objects.get(id=student_id)
                user.first_name = first_name
                user.last_name = last_name
                user.email = email
                user.username = username
                user.save()

                # Then Update Students Table
                student_model = Students.objects.get(admin=student_id)
                student_model.address = address

                gradelevel = GradeLevel.objects.get(id=GradeLevel_id)
                student_model.GradeLevel_id = gradelevel

                session_year_obj = SessionYearModel.objects.get(id=session_year_id)
                student_model.session_year_id = session_year_obj

                student_model.gender = gender
                if profile_pic_url != None:
                    student_model.profile_pic = profile_pic_url
                student_model.save()
                # Delete student_id SESSION after the data is updated
                del request.session['student_id']

                messages.success(request, "Student Updated Successfully!")
                return redirect('/edit_student/'+student_id)
            except:
                messages.success(request, "Failed to Uupdate Student.")
                return redirect('/edit_student/'+student_id)
        else:
            return redirect('/edit_student/'+student_id)


def deactivate_student(request, student_id):
    student = Students.objects.get(admin=student_id)
    try:
        # Set the is_active field of the associated CustomUser to False
        student.admin.is_active = False
        student.admin.save()
        
        messages.success(request, "Student Deactivated Successfully.")
        return redirect('manage_student')
    except Exception as e:
        print(e)
        messages.error(request, "Failed to Deactivate Student.")
        return redirect('manage_student')


def manage_section(request):
    gradelevels = GradeLevel.objects.all()
    sections = Section.objects.all()

    context = {
        "sections": sections,
        "gradelevels":gradelevels,
    }
    return render(request, "hod_template/Manage_Template/manage_section_template.html", context)

def add_section(request):
    gradelevels = GradeLevel.objects.all()
    context = {
        "gradelevels":gradelevels,
    }
    return render(request, 'hod_template/Add_Template/add_section_template.html', context)

def add_section_save(request):
    if request.method != "POST":
        messages.error(request, "Method Not Allowed!")
        return redirect('add_section')
    else:
        section_name = request.POST.get('section_name')
        section_limit = request.POST.get('section_limit')
        section_soft_limit = request.POST.get('section_soft_limit')
        GradeLevel_id = request.POST.get('gradelevel')
        gradelevel = GradeLevel.objects.get(id=GradeLevel_id)
        
        try:
            section = Section(
                section_name = section_name,
                section_soft_limit = section_soft_limit,
                section_limit = section_limit,
                GradeLevel_id=gradelevel)
            section.save()
            messages.success(request, "Section Added Successfully!")
            return redirect('add_section')
        except:
            messages.error(request, "Failed to Add Section!")
            return redirect('add_section')

def edit_section(request, section_id):

    section = Section.objects.get(id=section_id)
    gradelevel = GradeLevel.objects.all()

    context = {
        "section": section,
        "gradelevel":gradelevel,
    }

    return render(request, 'hod_template/Edit_Template/edit_section_template.html', context)

def edit_section_save(request):
    if request.method != "POST":
        HttpResponse("Invalid Method.")
    else:
        section_id = request.POST.get('section_id')
        section_name = request.POST.get('section_name')
        section_limit = request.POST.get('section_limit')
        section_soft_limit = request.POST.get('section_soft_limit')
        GradeLevel_id = request.POST.get('gradelevel')
        
        try:
            section = Section.objects.get(id=section_id)
            section.section_name = section_name

            gradelevel = GradeLevel.objects.get(id=GradeLevel_id)
            section.GradeLevel_id = gradelevel

            section_limit = section_limit
            section_soft_limit = section_soft_limit

            section.save()

            messages.success(request, "Section Updated Successfully.")
            # return redirect('/edit_subject/'+subject_id)
            return HttpResponseRedirect(reverse("edit_section", kwargs={"section_id":section_id}))
        except:
            messages.error(request, "Failed to Update Section.")
            return HttpResponseRedirect(reverse("edit_section", kwargs={"section_id":section_id}))
            # return redirect('/edit_subject/'+subject_id)


def add_subject(request):
    curriculums = Curriculums.objects.all()
    gradelevels = GradeLevel.objects.all()
    context = {
        "curriculums": curriculums,
        "gradelevels": gradelevels
        
    }
    return render(request, 'hod_template/Add_Template/add_subject_template.html', context)

def add_subject_save(request):
    if request.method != "POST":
        messages.error(request, "Method Not Allowed!")
        return redirect('add_subject')
    else:
        subject_name = request.POST.get('subject_name')
        subject_description = request.POST.get('subject_description')
        subject_code = request.POST.get('subject_code')
        subject_hours = request.POST.get('subject_hours')

        curriculum_id = request.POST.get('curriculum_id')
        curriculum_id = Curriculums.objects.get(id=curriculum_id)

        GradeLevel_id = request.POST.get('GradeLevel_id')
        GradeLevel_id = GradeLevel.objects.get(id=GradeLevel_id)
        
        try:
            subject = Subjects(subject_name=subject_name,
                               subject_description=subject_description,
                               subject_code=subject_code,
                               subject_hours=subject_hours,
                               GradeLevel_id=GradeLevel_id, 
                               curriculum_id = curriculum_id)
            subject.save()
            messages.success(request, "Subject Added Successfully!")
            return redirect('add_subject')
        except:
            messages.error(request, "Failed to Add Subject!")
            return redirect('add_subject')

def manage_subject(request):
    subjects = Subjects.objects.all()
    context = {
        "subjects": subjects
    }
    return render(request, 'hod_template/Manage_Template/manage_subject_template.html', context)

def edit_subject(request, subject_id):
    subject = Subjects.objects.get(id=subject_id)
    gradelevels = GradeLevel.objects.all()
    
    context = {
        "subject": subject,
        "gradelevels": gradelevels,
        "id": subject_id
    }
    return render(request, 'hod_template/Edit_Template/edit_subject_template.html', context)

def edit_subject_save(request):
    if request.method != "POST":
        HttpResponse("Invalid Method.")
    else:
        subject_id = request.POST.get('subject_id')
        GradeLevel_id = request.POST.get('gradelevel')
        subject_name = request.POST.get('subject')
        
        subject_code = request.POST.get('subject_code')
        subject_description = request.POST.get('subject_description')
        subject_hours = request.POST.get('subject_hours')
        
        try:
            subject = Subjects.objects.get(id=subject_id)
            subject.subject_name = subject_name
            
            gradelevel = GradeLevel.objects.get(id=GradeLevel_id)
            subject.GradeLevel_id = gradelevel
            
            subject.subject_code = subject_code
            subject.subject_description = subject_description
            subject.subject_hours = subject_hours
            
            subject.save()

            messages.success(request, "Subject Updated Successfully.")
            # return redirect('/edit_subject/'+subject_id)
            return HttpResponseRedirect(reverse("edit_subject", kwargs={"subject_id":subject_id}))

        except:
            messages.error(request, "Failed to Update Subject.")
            return HttpResponseRedirect(reverse("edit_subject", kwargs={"subject_id":subject_id}))
            # return redirect('/edit_subject/'+subject_id)

def delete_subject(request, subject_id):
    subject = Subjects.objects.get(id=subject_id)
    try:
        subject.delete()
        messages.success(request, "Subject Deleted Successfully.")
        return redirect('manage_subject')
    except:
        messages.error(request, "Failed to Delete Subject.")
        return redirect('manage_subject')
    
def add_assignsection(request):
    # Filter grade levels for students with student_status="Enrolled"
    gradelevels = GradeLevel.objects.filter(
        id__in=Students.objects.filter(
            student_status="Enrolled"  # Only include enrolled students
        ).values_list('GradeLevel_id', flat=True)
    )
    context = {
        "gradelevels": gradelevels,
    }
    return render(request, 'hod_template/Add_Template/add_assignsection_template.html', context)


# def add_assignsection(request):
#     # Filter grade levels only for students present in the Enrollment table
#     gradelevels = GradeLevel.objects.filter(
#         id__in=Students.objects.filter(
#             id__in=Enrollment.objects.values_list('student_id', flat=True)
#         ).values_list('GradeLevel_id', flat=True)
#     )
#     context = {
#         "gradelevels": gradelevels,
#     }
#     return render(request, 'hod_template/Add_Template/add_assignsection_template.html', context)

def load_sections_and_students(request):
    gradelevel_id = request.GET.get('gradelevel_id')
    
    # Fetch sections and students filtered by GradeLevel
    sections = Section.objects.filter(GradeLevel_id=gradelevel_id)

    # Get all students in the GradeLevel who are NOT already assigned to a section
    assigned_students = AssignSection.objects.filter(GradeLevel_id=gradelevel_id).values_list('Student_id', flat=True)
    students = Students.objects.filter(
        GradeLevel_id=gradelevel_id,
        student_status="Enrolled"  # Only include students with status "Enrolled"
    ).exclude(id__in=assigned_students)

    # Prepare data for response
    section_list = [{"id": section.id, "name": section.section_name} for section in sections]
    student_list = [{"id": student.id, "name": f"{student.admin.first_name} {student.admin.last_name}"} for student in students]
    
    return JsonResponse({'sections': section_list, 'students': student_list})

# def load_sections_and_students(request):
#     gradelevel_id = request.GET.get('gradelevel_id')
    
#     # Fetch sections by GradeLevel
#     sections = Section.objects.filter(GradeLevel_id=gradelevel_id)

#     # Get all students in the GradeLevel who are NOT already assigned to a section
#     assigned_students = AssignSection.objects.filter(GradeLevel_id=gradelevel_id).values_list('Student_id', flat=True)
#     students = Students.objects.filter(
#         GradeLevel_id=gradelevel_id,
#         id__in=Enrollment.objects.values_list('student_id', flat=True)  # Only include enrolled students
#     ).exclude(id__in=assigned_students)

#     # Prepare data for response
#     section_list = [{"id": section.id, "name": section.section_name} for section in sections]
#     student_list = [{"id": student.id, "name": f"{student.admin.first_name} {student.admin.last_name}"} for student in students]
    
#     return JsonResponse({'sections': section_list, 'students': student_list})


def add_assignsection_save(request):
    if request.method != "POST":
        messages.error(request, "Method Not Allowed!")
        return redirect('add_assignsection')
    else:
        try:
            gradelevel_id = request.POST.get('GradeLevel_id')
            gradelevel = GradeLevel.objects.get(id=gradelevel_id)

            section_id = request.POST.get('section_id')
            section = Section.objects.get(id=section_id)

            # Get the list of student_ids
            student_ids = request.POST.getlist('student_ids')

            if not student_ids:
                messages.error(request, "No students selected!")
                return redirect('add_assignsection')

            # Check the current count of students in this section
            current_count = AssignSection.objects.filter(section_id=section).count()
            remaining_slots = section.section_limit - current_count

            # Filter students to add based on remaining slots
            students_to_add = student_ids[:remaining_slots]  # Students that can be added within the limit
            students_not_added = student_ids[remaining_slots:]  # Students exceeding the limit

            # Proceed to save the assignment for each student within the limit
            for student_id in students_to_add:
                student = Students.objects.get(id=student_id)
                assignsection = AssignSection(
                    GradeLevel_id=gradelevel, 
                    section_id=section,
                    Student_id=student
                )
                assignsection.save()

            # Prepare a success message
            success_message = f"{len(students_to_add)} students assigned successfully to section '{section.section_name}'."
            
            # If there are students that couldn't be added, prepare a warning message
            if students_not_added:
                not_added_names = [Students.objects.get(id=student_id).admin.get_full_name() for student_id in students_not_added]
                warning_message = f"However, the following students could not be assigned due to the section limit of {section.section_limit}: {', '.join(not_added_names)}."
                messages.warning(request, warning_message)
            
            messages.success(request, success_message)
            return redirect('add_assignsection')

        except Exception as e:
            messages.error(request, f"Failed to Assign Section! Error: {e}")
            return redirect('add_assignsection')



def add_load(request):
    # Fetch all necessary data for the context
    session_years = SessionYearModel.objects.all()
    curriculums = Curriculums.objects.all()
    assignsections = AssignSection.objects.select_related('GradeLevel_id', 'section_id').distinct('GradeLevel_id', 'section_id')
    gradelevels = GradeLevel.objects.all()
    sections = Section.objects.all()
    subjects = Subjects.objects.all()
    
    staffs = CustomUser.objects.filter(user_type='2')
    # Fetching existing loads to display in the table
    loads = Load.objects.select_related('curriculum_id', 'AssignSection_id', 'subject_id', 'staff_id').all()  
    context = {
        "session_years":session_years,
        "curriculums": curriculums,
        "assignsections": assignsections,
        "gradelevels": gradelevels,
        "sections": sections,
        "subjects": subjects,
        "staffs": staffs,
        "loads": loads,
    }
    return render(request, 'hod_template/Add_Template/add_load_template.html', context)

def add_load_save(request):
    if request.method != "POST":
        messages.error(request, "Method Not Allowed!")
        return redirect('add_load')
    else:
        try:

            session_year_id = request.POST.get('session_year_id')
            session_id = SessionYearModel.objects.get(id=session_year_id)

            curriculum_id = request.POST.get('curriculum_id')
            curriculum = Curriculums.objects.get(id=curriculum_id)

            assignsection_id = request.POST.get('AssignSection_id')  # Adjusted to use AssignSection_id
            assignsection = AssignSection.objects.get(id=assignsection_id)

            subject_id = request.POST.get('subject_id')
            subject = Subjects.objects.get(id=subject_id)

            staff_id = request.POST.get('staff_id')
            staff_user = CustomUser.objects.get(id=staff_id)

            is_advisory = request.POST.get('is_advisory')

            # Check for duplicate load entry
            if Load.objects.filter(
                session_year_id=session_id,
                curriculum_id=curriculum,
                AssignSection_id=assignsection,
                subject_id=subject,
                staff_id=staff_user
            ).exists():
                messages.error(request, "This load configuration already exists.")
                return redirect('add_load')

            # Get the related Staff instance to check the max_load
            staff = Staffs.objects.get(admin=staff_user)
            current_load = Load.objects.filter(staff_id=staff_user).count()

            # Check if the current load exceeds or equals the staff's max_load
            if current_load >= staff.max_load:
                messages.error(request, f"{staff_user.first_name} {staff_user.last_name} has reached the maximum load limit of {staff.max_load}.")
                return redirect('add_load')

            # Saving the load
            load = Load(
                        session_year_id=session_id,      
                        curriculum_id=curriculum,
                        AssignSection_id=assignsection, 
                        subject_id=subject,
                        staff_id=staff_user,
                        is_advisory=is_advisory
                        )
            load.save()

            messages.success(request, "Load Added Successfully!")
            return redirect('add_load')

        except Exception as e:
            messages.error(request, f"Failed to Add Load! Error: {e}")
            return redirect('add_load')



def add_schedule(request):
    session_years = SessionYearModel.objects.all()
    staffs = CustomUser.objects.filter(user_type='2')
    loads = Load.objects.all()
    classrooms = Classroom.objects.all()
    context = {
        "session_years": session_years,
        "staffs": staffs,
        "loads": loads,  
        "classrooms": classrooms,
    }
    return render(request, 'hod_template/Add_Template/add_schedule_template.html', context)

def add_schedule_save(request):
    if request.method != "POST":
        messages.error(request, "Invalid Method")
        return redirect('add_schedule')
    else:
        
        try:

            session_year_id = request.POST.get('session_year_id')
            session_id = SessionYearModel.objects.get(id=session_year_id)
            
            staff_id = request.POST.get('staff_id')
            staff_id = CustomUser.objects.get(id=staff_id)

            load_id = request.POST.get('load_id')
            load_id = Load.objects.get(id=load_id)

            classroom_id = requests.POST.get('classroom_id')
            classroom_id = Classroom.objects.get(id=classroom_id)

            day_of_week = request.POST.get('day_of_week')
            start_time = request.POST.get('start_time')
            end_time = request.POST.get('end_time')
            
            # Saving the load
            schedule = Schedule(      
                        session_year_id=session_id,
                        staff_id=staff_id, 
                        load_id=load_id,
                        classroom_id=classroom_id,
                        day_of_week=day_of_week,
                        start_time=start_time,
                        end_time=end_time
                        )
            schedule.save()

            messages.success(request, "Schedule Added Successfully!")
            return redirect('add_schedule')
        
        except Exception as e:
            print(e)  # Print the exception for debugging purposes
            messages.error(request, f"Failed to Add Schedule! Erro: {e}" )
            return redirect('add_schedule')

def manage_class_scheduling(request):
    loads = Load.objects.all()
    assignsections = AssignSection.objects.all()
    schedules = Schedule.objects.all()
    classroom = Classroom.objects.all()
    context = {
        "loads": loads,
        "assignsections": assignsections,
        "schedules": schedules,
        "classroom": classroom,
    }
    return render(request, 'hod_template/Manage_Template/manage_schedule_template.html', context)


def class_search(request):
    search_query = request.GET.get('search', '')

    # Store search results
    assignsections = AssignSection.objects.all()
    loads = Load.objects.all()
    schedules = Schedule.objects.all()

    # If a search query is provided, filter the results
    if search_query:
        assignsections = AssignSection.objects.filter(
            Student_id__admin__first_name__icontains=search_query
        )
        loads = Load.objects.filter(
            subject_id__subject_name__icontains=search_query
        )
        schedules = Schedule.objects.filter(
            load_id__subject_id__subject_name__icontains=search_query
        )

        # Save the search query in the session to persist it across page loads
        request.session['search_query'] = search_query
    else:
        # Clear the search query from the session if no search is provided
        request.session.pop('search_query', None)

    # Check if there is a search query in the GET request and if a search has been made
    if 'search' in request.GET:
        # Don't redirect on the initial search
        return render(request, 'hod_template/Manage_Template/manage_schedule_template.html', {
            'assignsections': assignsections,
            'loads': loads,
            'schedules': schedules,
            'search_query': search_query,  # Pass the search query to the template if needed
        })

    # Render the search results or the default page
    return render(request, 'hod_template/Manage_Template/manage_schedule_template.html', {
        'assignsections': assignsections,
        'loads': loads,
        'schedules': schedules,
        'search_query': search_query,  # Pass the search query to the template if needed
    })


def edit_schedule(request, schedule_id):
    # Adding Schedule ID into Session Variable
    request.session['schedule_id'] = schedule_id

    try:
        # Retrieve the schedule object using the provided schedule_id
        schedule = Schedule.objects.get(id=schedule_id)
    except Schedule.DoesNotExist:
        messages.error(request, "Schedule does not exist.")
        return redirect('manage_class_scheduling')  # Adjust this redirect based on your URL names

    # Initialize the form with data from the database
    form = EditScheduleForm()
    form.fields['GradeLevel_id'].initial = schedule.GradeLevel_id.id
    form.fields['subject_id'].initial = schedule.subject_id.id
    form.fields['staff_id'].initial = schedule.staff_id.admin.id
    form.fields['session_year_id'].initial = schedule.session_year_id.id
    form.fields['day_of_week'].initial = schedule.day_of_week
    form.fields['start_time'].initial = schedule.start_time
    form.fields['end_time'].initial = schedule.end_time

    context = {
        "id": schedule_id,
        "form": form
    }

    return render(request, "hod_template/Edit_Template/edit_schedule_template.html", context)

def edit_schedule_save(request):
    if request.method != "POST":
        return HttpResponse("Invalid Method!")
    else:
        schedule_id = request.session.get('schedule_id')
        if schedule_id is None:
            return redirect('/manage_class_scheduling')

        form = EditScheduleForm(request.POST, request.FILES)
        if form.is_valid():
            subject_id = form.cleaned_data['subject_id']
            GradeLevel_id = form.cleaned_data['GradeLevel_id']
            staff_id = form.cleaned_data['staff_id']
            session_year_id = form.cleaned_data['session_year_id']
            day_of_week = form.cleaned_data['day_of_week']
            start_time = form.cleaned_data['start_time']
            end_time = form.cleaned_data['end_time']

            try:
                # Update Schedule model
                schedule = Schedule.objects.get(id=schedule_id)
                
                # Update schedule fields
                schedule.subject_id = Subjects.objects.get(subject_id)
                schedule.GradeLevel_id = GradeLevel.objects.get(id=GradeLevel_id)
                schedule.staff_id = Staffs.objects.get(id=staff_id)
                schedule.session_year_id = SessionYearModel.objects.get(id=session_year_id)
                schedule.day_of_week = day_of_week
                schedule.start_time = start_time
                schedule.end_time = end_time
                schedule.save()
                
                # Clear the session variable
                del request.session['schedule_id']

                messages.success(request, "Schedule Updated Successfully!")
                return redirect(f'/edit_schedule/{schedule_id}')
            except Exception as e:
                messages.error(request, f"Failed to Update Schedule: {e}")
                return redirect(f'/edit_schedule/{schedule_id}')
        else:
            return redirect(f'/edit_schedule/{schedule_id}')




@csrf_exempt
def check_email_exist(request):
    email = request.POST.get("email")
    user_obj = CustomUser.objects.filter(email=email).exists()
    if user_obj:
        return HttpResponse(True)
    else:
        return HttpResponse(False)


@csrf_exempt
def check_username_exist(request):
    username = request.POST.get("username")
    user_obj = CustomUser.objects.filter(username=username).exists()
    if user_obj:
        return HttpResponse(True)
    else:
        return HttpResponse(False)



def student_feedback_message(request):
    feedbacks = FeedBackStudent.objects.all()
    context = {
        "feedbacks": feedbacks
    }
    return render(request, 'hod_template/student_feedback_template.html', context)


@csrf_exempt
def student_feedback_message_reply(request):
    feedback_id = request.POST.get('id')
    feedback_reply = request.POST.get('reply')

    try:
        feedback = FeedBackStudent.objects.get(id=feedback_id)
        feedback.feedback_reply = feedback_reply
        feedback.save()
        return HttpResponse("True")

    except:
        return HttpResponse("False")


def staff_feedback_message(request):
    feedbacks = FeedBackStaffs.objects.all()
    context = {
        "feedbacks": feedbacks
    }
    return render(request, 'hod_template/staff_feedback_template.html', context)


@csrf_exempt
def staff_feedback_message_reply(request):
    feedback_id = request.POST.get('id')
    feedback_reply = request.POST.get('reply')

    try:
        feedback = FeedBackStaffs.objects.get(id=feedback_id)
        feedback.feedback_reply = feedback_reply
        feedback.save()
        return HttpResponse("True")

    except:
        return HttpResponse("False")


def student_leave_view(request):
    leaves = LeaveReportStudent.objects.all()
    context = {
        "leaves": leaves
    }
    return render(request, 'hod_template/student_leave_view.html', context)

def student_leave_approve(request, leave_id):
    leave = LeaveReportStudent.objects.get(id=leave_id)
    leave.leave_status = 1
    leave.save()
    return redirect('student_leave_view')


def student_leave_reject(request, leave_id):
    leave = LeaveReportStudent.objects.get(id=leave_id)
    leave.leave_status = 2
    leave.save()
    return redirect('student_leave_view')


def staff_leave_view(request):
    leaves = LeaveReportStaff.objects.all()
    context = {
        "leaves": leaves
    }
    return render(request, 'hod_template/staff_leave_view.html', context)


def staff_leave_approve(request, leave_id):
    leave = LeaveReportStaff.objects.get(id=leave_id)
    leave.leave_status = 1
    leave.save()
    return redirect('staff_leave_view')


def staff_leave_reject(request, leave_id):
    leave = LeaveReportStaff.objects.get(id=leave_id)
    leave.leave_status = 2
    leave.save()
    return redirect('staff_leave_view')


def admin_view_attendance(request):
    subjects = Subjects.objects.all()
    session_years = SessionYearModel.objects.all()
    context = {
        "subjects": subjects,
        "session_years": session_years
    }
    return render(request, "hod_template/admin_view_attendance.html", context)


@csrf_exempt
def admin_get_attendance_dates(request):
    # Getting Values from Ajax POST 'Fetch Student'
    subject_id = request.POST.get("subject")
    session_year = request.POST.get("session_year_id")

    # Students enroll to GradeLevel, GradeLevel has Subjects
    # Getting all data from subject model based on subject_id
    subject_model = Subjects.objects.get(id=subject_id)

    session_model = SessionYearModel.objects.get(id=session_year)

    # students = Students.objects.filter(GradeLevel_id=subject_model.GradeLevel_id, session_year_id=session_model)
    attendance = Attendance.objects.filter(subject_id=subject_model, session_year_id=session_model)

    # Only Passing Student Id and Student Name Only
    list_data = []

    for attendance_single in attendance:
        data_small={"id":attendance_single.id, "attendance_date":str(attendance_single.attendance_date), "session_year_id":attendance_single.session_year_id.id}
        list_data.append(data_small)

    return JsonResponse(json.dumps(list_data), content_type="application/json", safe=False)


@csrf_exempt
def admin_get_attendance_student(request):
    # Getting Values from Ajax POST 'Fetch Student'
    attendance_date = request.POST.get('attendance_date')
    attendance = Attendance.objects.get(id=attendance_date)

    attendance_data = AttendanceReport.objects.filter(attendance_id=attendance)
    # Only Passing Student Id and Student Name Only
    list_data = []

    for student in attendance_data:
        data_small={"id":student.student_id.admin.id, "name":student.student_id.admin.first_name+" "+student.student_id.admin.last_name, "status":student.status}
        list_data.append(data_small)

    return JsonResponse(json.dumps(list_data), content_type="application/json", safe=False)


def admin_profile(request):
    user = CustomUser.objects.get(id=request.user.id)

    context={
        "user": user
    }
    return render(request, 'hod_template/admin_profile.html', context)


def admin_profile_update(request):
    if request.method != "POST":
        messages.error(request, "Invalid Method!")
        return redirect('admin_profile')
    else:
        first_name = request.POST.get('first_name')
        last_name = request.POST.get('last_name')
        password = request.POST.get('password')

        try:
            customuser = CustomUser.objects.get(id=request.user.id)
            customuser.first_name = first_name
            customuser.last_name = last_name
            if password != None and password != "":
                customuser.set_password(password)
            customuser.save()
            messages.success(request, "Profile Updated Successfully")
            return redirect('admin_profile')
        except:
            messages.error(request, "Failed to Update Profile")
            return redirect('admin_profile')
    


def staff_profile(request):
    pass


def student_profile(requtest):
    pass



