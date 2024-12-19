from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment
from django.utils.timezone import now
from django.contrib.auth.hashers import make_password
from decimal import Decimal, InvalidOperation
from django.db import IntegrityError
from django.db.models import Count, F
from django.db import transaction
from datetime import datetime, date
from bs4 import BeautifulSoup
from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse, HttpResponseRedirect, JsonResponse
from django.contrib import messages
from django.urls import reverse
from django.views.decorators.csrf import csrf_exempt
import json
from datetime import datetime, date, time
from datetime import datetime as dt
from django.utils import timezone
from django.db.models import Q, Subquery
import cloudinary.uploader
import openpyxl
import requests
from django.contrib.auth.decorators import login_required

from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.exceptions import AuthenticationFailed
from rest_framework_simplejwt.tokens import RefreshToken
from rest_framework.permissions import IsAuthenticated
from student_management_app.EmailBackEnd import EmailBackEnd
from rest_framework_simplejwt.exceptions import TokenError
from rest_framework.permissions import AllowAny
from django.contrib.auth import get_user_model

from student_management_app.models import CustomUser, School_info, AdminHOD, Students, ParentGuardian, PreviousSchool, EmergencyContact, Attachment, BalancePayment, AssignSection, Load, Schedule, GradingConfiguration
from student_management_app.models import Staffs, staff_contact_info, staff_employment_info, staff_physical_info, staff_government_ID_info, Staffs_Educ_Background, StudentPromotionHistory
from student_management_app.models import Curriculums, GradeLevel, Enrollment, Enrollment_voucher, Subjects, Section, SessionYearModel
from student_management_app.models import FeedBackStudent, FeedBackStaffs, LeaveReportStudent, LeaveReportStaff, Attendance, AttendanceReport, Classroom, StudentResult
from .forms import EditStudentForm, AddScheduleForm, EditScheduleForm


def admin_home(request):
    all_student_count = Students.objects.count()
    subject_count = Subjects.objects.count()
    gradelevel_count = GradeLevel.objects.count()
    staff_count = Staffs.objects.count()
    load_count = Load.objects.count()
    curriculum_count = Curriculums.objects.count()

    # Serialize GradeLevel data
    grade_levels = GradeLevel.objects.values('GradeLevel_name')
    grade_levels_data = json.dumps(list(grade_levels))  # Serialize data for JSON compatibility

    # Updated query to get subjects grouped by GradeLevel
    subjects_per_gradelevel = Subjects.objects.select_related('GradeLevel_id') \
        .values('GradeLevel_id__GradeLevel_name') \
        .annotate(subject_count=Count('id'))

    # Prepare data to be passed to Chart.js
    gradelevels_datas = [
        {
            "grade_level": item['GradeLevel_id__GradeLevel_name'],  # The grade level name
            "subject_count": item['subject_count']  # The count of subjects per grade level
        }
        for item in subjects_per_gradelevel
    ]

    # Aggregate the number of students per GradeLevel
    students_per_gradelevel = Students.objects.values('GradeLevel_id__GradeLevel_name').annotate(student_count=Count('id'))
    
    # Prepare the data to pass to the template
    student_gradelevel_data = [
        {
            "grade_level": item['GradeLevel_id__GradeLevel_name'],
            "student_count": item['student_count']
        }
        for item in students_per_gradelevel
    ]

    context = {
        "all_student_count": all_student_count,
        "subject_count": subject_count,
        "gradelevel_count": gradelevel_count,
        "staff_count": staff_count,
        "load_count": load_count,
        "curriculum_count": curriculum_count,
        "curriculum_count": curriculum_count,
        "grade_levels_data": grade_levels_data,  # Pass JSON-serialized data to context
        "gradelevels_datas": json.dumps(gradelevels_datas),  # Pass JSON-serialized data to context
        "student_gradelevel_data": json.dumps(student_gradelevel_data),  # Pass JSON-serialized data to context
    }

    return render(request, "hod_template/home_content.html", context)

def add_staff(request):
    return render(request, "hod_template/Add_Template/add_staff_template.html")

def add_admin(request):
    return render(request, "hod_template/Add_Template/add_admin_template.html")

def add_staff_save(request):
    if request.method != "POST":
        messages.error(request, "Invalid Method ")
        return redirect('add_staff')
    else:
        try:
            first_name = request.POST.get('first_name')
            last_name = request.POST.get('last_name')
            username = request.POST.get('username')
            email = request.POST.get('email')
            middle_name = request.POST.get('middle_name')
            dob = request.POST.get('dob')
            age = request.POST.get('age')
            pob = request.POST.get('pob')
            sex = request.POST.get('sex')
            civil_status = request.POST.get('civil_status')
            citizenship = request.POST.get('citizenship')
            dual_country = request.POST.get('dual_country')
            max_load = request.POST.get('max_load')

            # Generate a predefined password based on staff details
            current_year = now().year
            predefined_password = f"{first_name.lower()}{last_name.lower()}{current_year}"

            # Create Staff User
            user = CustomUser.objects.create_user(
                username=username,
                password=predefined_password,  # Use hashed password
                email=email,
                first_name=first_name,
                last_name=last_name,
                user_type=2
            )
            user.staffs.middle_name = middle_name
            user.staffs.dob = dob
            user.staffs.age = age
            user.staffs.pob = pob
            user.staffs.sex = sex
            user.staffs.civil_status = civil_status
            user.staffs.citizenship = citizenship
            user.staffs.dual_country = dual_country
            user.staffs.max_load = max_load
            user.save()

            # save staff contact information
            staff_contact_info.objects.create(
                 staffs_id = user.staffs,
                 region = request.POST.get('region-text'),
                 province = request.POST.get('province-text'),
                 city = request.POST.get('city-text'),
                 barangay = request.POST.get('barangay-text'),
                 street = request.POST.get('street'),
                 telephone_no = request.POST.get('telephone_no'),
                 cellphone_no = request.POST.get('cellphone_no'),
                 emergency_contact_name = request.POST.get('emergency_contact_name'),
                 emergency_contact_no = request.POST.get('emergency_contact_no'),
                 emergency_relationship = request.POST.get('emergency_relationship'),
                 medical_condition = request.POST.get('medical_condition'),
            )
            # save staff employment information
            staff_employment_info.objects.create(
                 staffs_id = user.staffs,
                 employee_number = request.POST.get('employee_number'),
                 employee_type = request.POST.get('employee_type'),
                 position = request.POST.get('position'),
                 employment_status = request.POST.get('employment_status'),
            )
            # save staff physical information
            staff_physical_info.objects.create(
                 staffs_id = user.staffs,
                 blood_type = request.POST.get('blood_type'),
                 height = request.POST.get('height'),
                 weight = request.POST.get('weight'),
                 eye_color = request.POST.get('eye_color'),
                 hair_color = request.POST.get('hair_color'),
            )
            # save staff govenrment ID information
            staff_government_ID_info.objects.create(
                 staffs_id = user.staffs,
                 gsis_id = request.POST.get('gsis_id'),
                 philhealth_id = request.POST.get('philhealth_id'),
                 pagibig_id = request.POST.get('pagibig_id'),
                 sss_id = request.POST.get('sss_id'),
                 tin_id = request.POST.get('tin_id'),
            )
            # save staff educational background information
            Staffs_Educ_Background.objects.create(
                 staffs_id = user.staffs,
                 HEA = request.POST.get('HEA'),
                 preferred_subject = request.POST.get('preferred_subject'),
                 Cert_License = request.POST.get('Cert_License'),
                 teaching_exp = request.POST.get('teaching_exp'),
                 skills_competencies = request.POST.get('skills_competencies'),
                 language_spoken = request.POST.get('language_spoken'),
            )
            messages.success(request, "Faculty Added Successfully! Password: " + predefined_password)
            return redirect('add_staff')
        
        except IntegrityError as e:
            messages.error(request, f"Failed to Add Staff: Database error - {str(e)}")
            return redirect('add_staff')
        except Exception as e:
            messages.error(request, f"Failed to Add Staff: {str(e)}")
            return redirect('add_staff')
 
def add_admin_save(request):
    if request.method != "POST":
        messages.error(request, "Invalid Method ")
        return redirect('add_admin')
    else:
        try:
            first_name = request.POST.get('first_name')
            last_name = request.POST.get('last_name')
            username = request.POST.get('username')
            email = request.POST.get('email')
            middle_name = request.POST.get('middle_name')
            dob = request.POST.get('dob')
            age = request.POST.get('age')
            pob = request.POST.get('pob')
            sex = request.POST.get('sex')
            civil_status = request.POST.get('civil_status')
            citizenship = request.POST.get('citizenship')
            dual_country = request.POST.get('dual_country')
            max_load = request.POST.get('max_load')

            # Generate a predefined password based on staff details
            current_year = now().year
            predefined_password = f"{first_name.lower()}{last_name.lower()}{current_year}"

            # Create Staff User
            user = CustomUser.objects.create_user(
                username=username,
                password=predefined_password,  # Use hashed password
                email=email,
                first_name=first_name,
                last_name=last_name,
                user_type=1
            )
            user.adminhod.middle_name = middle_name
            user.adminhod.dob = dob
            user.adminhod.age = age
            user.adminhod.pob = pob
            user.adminhod.sex = sex
            user.adminhod.civil_status = civil_status
            user.adminhod.citizenship = citizenship
            user.adminhod.dual_country = dual_country
            user.adminhod.max_load = max_load
            user.save()

            # save staff contact information
            staff_contact_info.objects.create(
                 adminhod_id = user.adminhod,
                 region = request.POST.get('region-text'),
                 province = request.POST.get('province-text'),
                 city = request.POST.get('city-text'),
                 barangay = request.POST.get('barangay-text'),
                 street = request.POST.get('street'),
                 telephone_no = request.POST.get('telephone_no'),
                 cellphone_no = request.POST.get('cellphone_no'),
                 emergency_contact_name = request.POST.get('emergency_contact_name'),
                 emergency_contact_no = request.POST.get('emergency_contact_no'),
                 emergency_relationship = request.POST.get('emergency_relationship'),
                 medical_condition = request.POST.get('medical_condition'),
            )
            # save staff employment information
            staff_employment_info.objects.create(
                 adminhod_id = user.adminhod,
                 employee_number = request.POST.get('employee_number'),
                 employee_type = request.POST.get('employee_type'),
                 position = request.POST.get('position'),
                 employment_status = request.POST.get('employment_status'),
            )
            # save staff physical information
            staff_physical_info.objects.create(
                 adminhod_id = user.adminhod,
                 blood_type = request.POST.get('blood_type'),
                 height = request.POST.get('height'),
                 weight = request.POST.get('weight'),
                 eye_color = request.POST.get('eye_color'),
                 hair_color = request.POST.get('hair_color'),
            )
            # save staff govenrment ID information
            staff_government_ID_info.objects.create(
                 adminhod_id = user.adminhod,
                 gsis_id = request.POST.get('gsis_id'),
                 philhealth_id = request.POST.get('philhealth_id'),
                 pagibig_id = request.POST.get('pagibig_id'),
                 sss_id = request.POST.get('sss_id'),
                 tin_id = request.POST.get('tin_id'),
            )
            # save staff educational background information
            Staffs_Educ_Background.objects.create(
                 adminhod_id = user.adminhod,
                 HEA = request.POST.get('HEA'),
                 preferred_subject = request.POST.get('preferred_subject'),
                 Cert_License = request.POST.get('Cert_License'),
                 teaching_exp = request.POST.get('teaching_exp'),
                 skills_competencies = request.POST.get('skills_competencies'),
                 language_spoken = request.POST.get('language_spoken'),
            )
            messages.success(request, "Admin Added Successfully! Password: " + predefined_password)
            return redirect('add_admin')
        
        except IntegrityError as e:
            messages.error(request, f"Failed to Add Admin: Database error - {str(e)}")
            return redirect('add_admin')
        except Exception as e:
            messages.error(request, f"Failed to Add Admin: {str(e)}")
            return redirect('add_admin')

def toggle_user_activation(request, user_id):
    user = get_user_model().objects.get(id=user_id)

    if user.is_active:
        user.is_active = False
        messages.success(request, f"User {user.username} has been deactivated.")
    else:
        user.is_active = True
        messages.success(request, f"User {user.username} has been reactivated.")
    
    user.save()

    return redirect('manage_staff')

def toggle_grading_state(request):
    grading_config, created = GradingConfiguration.objects.get_or_create(id=1)
    # Toggle the grading state
    grading_config.is_grading_active = not grading_config.is_grading_active
    grading_config.save()
    
    # Redirect back to the manage_staff view
    return redirect('manage_staff')

def manage_staff(request):
    # Retrieve all AdminHOD and Staff instances
    admins = AdminHOD.objects.all()
    staffs = Staffs.objects.all()
    
    # Initialize an empty list to hold combined user details
    all_users = []

    # Retrieve and structure data for AdminHODs
    for admin in admins:
        # Retrieve associated user data
        admin_data = {
            'user': admin.admin,
            'user_id': admin.admin.id,
            'first_name': admin.admin.first_name,  
            'last_name': admin.admin.last_name,    
            'username': admin.admin.username,      
            'email': admin.admin.email,            
            'last_login': admin.admin.last_login,  
            'date_joined': admin.admin.date_joined, 
            'user_type': admin.admin.user_type,   
            'middle_name': admin.middle_name,
            'suffix': admin.suffix,
            'dob': admin.dob,
            'age': admin.age,
            'pob': admin.pob,
            'sex': admin.sex,
            'civil_status': admin.civil_status,
            'citizenship': admin.citizenship,
            'dual_country': admin.dual_country,
            'max_load': admin.max_load,
            'contact_info': staff_contact_info.objects.filter(adminhod_id=admin).first(),
            'employment_info': staff_employment_info.objects.filter(adminhod_id=admin).first(),
            'physical_info': staff_physical_info.objects.filter(adminhod_id=admin).first(),
            'gov_id_info': staff_government_ID_info.objects.filter(adminhod_id=admin).first(),
            'education_info': Staffs_Educ_Background.objects.filter(adminhod_id=admin).first()
        }
        all_users.append(admin_data)

    # Retrieve and structure data for Staffs
    for staff in staffs:
        # Retrieve associated user data
        staff_data = {
            'user': staff.admin,
            'user_id': staff.admin.id,
            'first_name': staff.admin.first_name,  
            'last_name': staff.admin.last_name,    
            'username': staff.admin.username,      
            'email': staff.admin.email,            
            'last_login': staff.admin.last_login,  
            'date_joined': staff.admin.date_joined, 
            'user_type': staff.admin.user_type,   
            'middle_name': staff.middle_name,
            'suffix': staff.suffix,
            'dob': staff.dob,
            'age': staff.age,
            'pob': staff.pob,
            'sex': staff.sex,
            'civil_status': staff.civil_status,
            'citizenship': staff.citizenship,
            'dual_country': staff.dual_country,
            'max_load': staff.max_load,
            'contact_info': staff_contact_info.objects.filter(staffs_id=staff).first(),
            'employment_info': staff_employment_info.objects.filter(staffs_id=staff).first(),
            'physical_info': staff_physical_info.objects.filter(staffs_id=staff).first(),
            'gov_id_info': staff_government_ID_info.objects.filter(staffs_id=staff).first(),
            'education_info': Staffs_Educ_Background.objects.filter(staffs_id=staff).first()
        }
        all_users.append(staff_data)

    # Grading configuration (as before)
    grading_config, created = GradingConfiguration.objects.get_or_create(id=1)

    # Context to pass to the template
    context = {
        'all_users': all_users,
        'grading_config': grading_config
    }

    # Render the template with the context
    return render(request, "hod_template/Manage_Template/manage_staff_template.html", context)

def edit_staff(request, staff_id):
    # Retrieve the staff object using the staff_id directly
    staff = get_object_or_404(Staffs, id=staff_id)  # Use id, not admin
    # Retrieve associated information
    staffContactInfo = staff_contact_info.objects.filter(staffs_id=staff).first()
    staffEmploymentInfo = staff_employment_info.objects.filter(staffs_id=staff).first()
    staffPhysicalInfo = staff_physical_info.objects.filter(staffs_id=staff).first()
    staffGovernmentId = staff_government_ID_info.objects.filter(staffs_id=staff).first()
    staffEducBackground = Staffs_Educ_Background.objects.filter(staffs_id=staff).first()

    # Pass data to the template
    context = {
        "staff": staff,
        'first_name': staff.admin.first_name,  
        'last_name': staff.admin.last_name,    
        'username': staff.admin.username,      
        'email': staff.admin.email,    
        "id": staff_id,
        "staffContactInfo": staffContactInfo,
        "staffEmploymentInfo": staffEmploymentInfo,
        "staffPhysicalInfo": staffPhysicalInfo,
        "staffGovernmentId": staffGovernmentId,
        "staffEducBackground": staffEducBackground,
        "blood_type": staffPhysicalInfo.blood_type,
    }

    return render(request, "hod_template/Edit_Template/edit_staff_template.html", context)

def edit_staff_save(request, staff_id):
    if request.method != "POST":
        messages.error(request, "Method Not Allowed!")
        return redirect('edit_staff', staff_id=staff_id)

    try:
        # Fetch the staff based on the provided staff_id
        staff = Staffs.objects.get(admin=staff_id)
        user = staff.admin  # The associated CustomUser instance (admin)

        # Extract data from request.POST
        first_name = request.POST.get('first_name')
        last_name = request.POST.get('last_name')
        username = request.POST.get('username')
        email = request.POST.get('email')
        middle_name = request.POST.get('middle_name')
        dob = request.POST.get('dob')
        age = request.POST.get('age')
        pob = request.POST.get('pob')
        sex = request.POST.get('sex')
        civil_status = request.POST.get('civil_status')
        citizenship = request.POST.get('citizenship')
        dual_country = request.POST.get('dual_country')
        max_load = request.POST.get('max_load')

        # Update user details
        user.username = username
        user.email = email
        user.first_name = first_name
        user.last_name = last_name
        user.save()  # Save the CustomUser instance

        # Update staff details
        staff.middle_name = middle_name
        staff.dob = dob
        staff.age = age
        staff.pob = pob
        staff.sex = sex
        staff.civil_status = civil_status
        staff.citizenship = citizenship
        staff.dual_country = dual_country
        staff.max_load = max_load
        staff.save()  # Save the Staff instance

        # Update Staff Contact Information
        contact_info, _ = staff_contact_info.objects.get_or_create(staffs_id=staff)
        contact_info.region = request.POST.get('region-text')
        contact_info.province = request.POST.get('province-text')
        contact_info.city = request.POST.get('city-text')
        contact_info.barangay = request.POST.get('barangay-text')
        contact_info.street = request.POST.get('street')
        contact_info.telephone_no = request.POST.get('telephone_no')
        contact_info.cellphone_no = request.POST.get('cellphone_no')
        contact_info.emergency_contact_name = request.POST.get('emergency_contact_name')
        contact_info.emergency_contact_no = request.POST.get('emergency_contact_no')
        contact_info.emergency_relationship = request.POST.get('emergency_relationship')
        contact_info.medical_condition = request.POST.get('medical_condition')
        contact_info.save()

        # Update Staff Employment Information
        employment_info, _ = staff_employment_info.objects.get_or_create(staffs_id=staff)
        employment_info.employee_number = request.POST.get('employee_number')
        employment_info.employee_type = request.POST.get('employee_type')
        employment_info.position = request.POST.get('position')
        employment_info.employment_status = request.POST.get('employment_status')
        employment_info.save()

        # Update Staff Physical Information
        physical_info, _ = staff_physical_info.objects.get_or_create(staffs_id=staff)
        physical_info.blood_type = request.POST.get('blood_type')
        physical_info.height = request.POST.get('height')
        physical_info.weight = request.POST.get('weight')
        physical_info.eye_color = request.POST.get('eye_color')
        physical_info.hair_color = request.POST.get('hair_color')
        physical_info.save()

        # Update Staff Government ID Information
        government_info, _ = staff_government_ID_info.objects.get_or_create(staffs_id=staff)
        government_info.gsis_id = request.POST.get('gsis_id')
        government_info.philhealth_id = request.POST.get('philhealth_id')
        government_info.pagibig_id = request.POST.get('pagibig_id')
        government_info.sss_id = request.POST.get('sss_id')
        government_info.tin_id = request.POST.get('tin_id')
        government_info.save()

        # Update Staff Educational Background Information
        education_info, _ = Staffs_Educ_Background.objects.get_or_create(staffs_id=staff)
        education_info.HEA = request.POST.get('HEA')
        education_info.preferred_subject = request.POST.get('preferred_subject')
        education_info.Cert_License = request.POST.get('Cert_License')
        education_info.teaching_exp = request.POST.get('teaching_exp')
        education_info.skills_competencies = request.POST.get('skills_competencies')
        education_info.language_spoken = request.POST.get('language_spoken')
        education_info.save()

        # Success message
        messages.success(request, "Staff Information Updated Successfully!")
        return redirect('edit_staff', staff_id=staff_id)

    except Staffs.DoesNotExist:
        messages.error(request, "Staff Not Found!")
        return redirect('edit_staff', staff_id=staff_id)

    except Exception as e:
        messages.error(request, f"Failed to Update Staff! Error: {str(e)}")
        return redirect('edit_staff', staff_id=staff_id)

def deactivate_staff(request, staff_id):
    try:
        # Get the CustomUser object related to the staff
        user = CustomUser.objects.get(id=staff_id)

        # Set the user as inactive
        user.is_active = False
        user.save()

        messages.success(request, "Staff account deactivated successfully.")
        return redirect('manage_staff')
    except CustomUser.DoesNotExist:
        messages.error(request, "Staff not found.")
        return redirect('manage_staff')
    except Exception as e:
        messages.error(request, f"Failed to deactivate staff: {str(e)}")
        return redirect('manage_staff')

def add_school(request):
    return render(request, "hod_template/Add_Template/add_school_template.html")

def add_school_save(request):
    if request.method != "POST":
        messages.error(request, "Invalid Method!")
        return redirect('add_school')
    else:
        school_name = request.POST.get('school_name')
        school_ID_number = request.POST.get('school_ID_number')
        school_district = request.POST.get('school_district')
        school_division = request.POST.get('school_division')
        school_region = request.POST.get('school_region')
        region = request.POST.get('region-text')
        province = request.POST.get('province-text')
        city = request.POST.get('city-text')
        barangay = request.POST.get('barangay-text')
        street = request.POST.get('street')
        school_email = request.POST.get('school_email')
        school_cellphone = request.POST.get('school_cellphone')
        school_telephone = request.POST.get('school_telephone')
        try:
            school = School_info(
                school_name=school_name,
                school_ID_number=school_ID_number,
                school_district=school_district,
                school_division=school_division,
                school_region=school_region,
                region=region,
                province=province,
                city=city,
                barangay=barangay,
                street=street,
                school_email=school_email,
                school_cellphone=school_cellphone,
                school_telephone=school_telephone
                )
            school.save()
            messages.success(request, "School Added Successfully!")
            return redirect('add_school')
        except:
            messages.error(request, "Failed to Add School!")
            return redirect('add_school')
        
def manage_school(request):
    schools = School_info.objects.all()
    context = {
        "schools": schools
    }
    return render(request, 'hod_template/Manage_Template/manage_school_template.html', context)

def edit_school(request, school_info_id):
    school = School_info.objects.get(id=school_info_id)
    context = {
        "school": school
    }
    return render(request, 'hod_template/Edit_Template/edit_school_info_template.html', context)

def edit_school_save(request):
    if request.method == "POST":
        school_id = request.POST.get('school_id')
        school_name = request.POST.get('school_name')
        school_ID_number = request.POST.get('school_ID_number')
        region = request.POST.get('region')
        province = request.POST.get('province')
        city = request.POST.get('city')
        barangay = request.POST.get('barangay')
        street = request.POST.get('street')

        try:
            school = School_info.objects.get(id=school_id)
            school.school_name = school_name
            school.school_ID_number = school_ID_number
            school.region = region
            school.province = province
            school.city = city
            school.barangay = barangay
            school.street = street
            school.save()

            messages.success(request, "School Information Updated Successfully.")
            return redirect('/manage_school/')
        except Exception as e:
            messages.error(request, f"Failed to Update School Information. Error: {e}")
            return redirect('/manage_school/')
    else:
        return HttpResponse("Invalid Method")

def add_curriculum(request):
    return render(request, "hod_template/Add_Template/add_curriculum_template.html")


def add_curriculum_save(request):
    if request.method != "POST":
        messages.error(request, "Invalid Method!")
        return redirect('add_curriculum')
    else:
        curriculum_name = request.POST.get('curriculum_name')
        curriculum_description = request.POST.get('curriculum_description')
        curriculum_status = request.POST.get('curriculum_status')
        try:
            curriculum = Curriculums(
                curriculum_name=curriculum_name,
                curriculum_description=curriculum_description,
                curriculum_status=curriculum_status
                )
            curriculum.save()
            messages.success(request, "Curriculum Added Successfully!")
            return redirect('add_curriculum')
        except:
            messages.error(request, "Failed to Add Curriculum!")
            return redirect('add_curriculum')
        

def manage_curriculum(request):
    curriculums = Curriculums.objects.all()
    curriculums = Curriculums.objects.filter(is_archived=False)
    context = {
        "curriculums": curriculums
    }
    return render(request, 'hod_template/Manage_Template/manage_curriculum_template.html', context)

def edit_curriculum(request, curriculum_id):
    curriculum = Curriculums.objects.get(id=curriculum_id)
    context = {
        "curriculum": curriculum,
        "id": curriculum_id
    }
    return render(request, 'hod_template/Edit_Template/edit_curriculum_template.html', context)

def edit_curriculum_save(request):
    if request.method != "POST":
        HttpResponse("Invalid Method")
    else:
        curriculum_id = request.POST.get('curriculum_id')
        curriculum_name = request.POST.get('curriculum_name')
        curriculum_description = request.POST.get('curriculum_description')
        curriculum_status = request.POST.get('curriculum_status')
        try:
            curriculum = Curriculums.objects.get(id=curriculum_id)
            curriculum.curriculum_name = curriculum_name
            curriculum.curriculum_description = curriculum_description
            curriculum.curriculum_status = curriculum_status
            curriculum.save()
            messages.success(request, "Curriculum Updated Successfully.")
            return redirect('/manage_curriculum/')
        except:
            messages.error(request, "Failed to Update GradeLevel.")
            return redirect('/manage_curriculum/')
def archived_curriculums(request):
    # Filter curriculums with status 'Archived'
    curriculums = Curriculums.objects.filter(is_archived=True)
    context = {
        "curriculums": curriculums
    }
    return render(request, 'hod_template/Archive_Template/archived_curriculum_template.html', context)

def archive_curriculum(request, curriculum_id):
    curriculum = get_object_or_404(Curriculums, id=curriculum_id)
    curriculum.is_archived = True
    curriculum.curriculum_status = "Inactive"
    curriculum.save()
    messages.success(request, "Curriculum archived successfully!")
    return redirect('/manage_curriculum/')  
def unarchive_curriculum(request, curriculum_id):
    curriculum = get_object_or_404(Curriculums, id=curriculum_id)
    curriculum.is_archived = False
    curriculum.curriculum_status = "Active"
    curriculum.save()
    messages.success(request, "Curriculum unarchived successfully!")
    return redirect('/manage_curriculum/')  


def add_gradelevel(request):
    curriculums = Curriculums.objects.all()
    context = {
        "curriculums": curriculums
    }
    return render(request, "hod_template/Add_Template/add_gradelevel_template.html", context)

def add_gradelevel_save(request):
    if request.method != "POST":
        messages.error(request, "Invalid Method!")
        return redirect('add_gradelevel')
    else:

        GradeLevel_name = request.POST.get('GradeLevel_name')

        curriculum_id = request.POST.get('curriculum_id')
        curriculums = Curriculums.objects.get(id=curriculum_id)

        try:
            gradelevel = GradeLevel(
                GradeLevel_name=GradeLevel_name,
                curriculum_id = curriculums
                )
            gradelevel.save()
            messages.success(request, "GradeLevel Added Successfully!")
            return redirect('add_gradelevel')
        except:
            messages.error(request, "Failed to Add GradeLevel!")
            return redirect('add_gradelevel')
        
def export_gradelevels_to_excel(request):
    # Create an in-memory workbook
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = 'Grade Levels'
    
    # Define the headers
    headers = ['GradeLevel Name', 'Curriculum', 'Date Created', 'Date Updated']
    
    # Define a fill color for the header (e.g., light blue)
    header_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
    
    # Add headers to the first row and make them bold with a background color and center alignment
    for col_num, header in enumerate(headers, 1):
        cell = sheet.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)  # Make the header bold
        cell.fill = header_fill  # Apply background color
        cell.alignment = Alignment(horizontal="center", vertical="center")  # Center-align the text

    # Query all grade levels
    gradelevels = GradeLevel.objects.all()

    # Add data rows starting from row 2 with center alignment
    for row_num, gradelevel in enumerate(gradelevels, 2):
        sheet.cell(row=row_num, column=1).value = gradelevel.GradeLevel_name
        sheet.cell(row=row_num, column=2).value = gradelevel.curriculum_id.curriculum_name  # Assuming curriculum has 'curriculum_name'
        sheet.cell(row=row_num, column=3).value = gradelevel.created_at.strftime('%Y-%m-%d %H:%M:%S')
        sheet.cell(row=row_num, column=4).value = gradelevel.updated_at.strftime('%Y-%m-%d %H:%M:%S')

        # Apply center alignment to all cells in the row
        for col_num in range(1, 5):
            cell = sheet.cell(row=row_num, column=col_num)
            cell.alignment = Alignment(horizontal="center", vertical="center")

    # Adjust column widths to fit the content
    for col in sheet.columns:
        max_length = 0
        column = col[0].column_letter  # Get the column letter (A, B, C, etc.)
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = max_length + 2  # Add some padding to the width
        sheet.column_dimensions[column].width = adjusted_width

    # Prepare the response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=GradeLevels.xlsx'

    # Save the workbook to the response
    workbook.save(response)

    return response

def export_staffs_to_excel(request):
    # Create an in-memory workbook
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = 'Staffs'
    
    # Define the headers
    headers = ['Last Name', 'First Name', 'Middle Name', 'Email', 'Age', 'Date of Birth', 'Place of Birth', 
               'Sex', 'Civil Rights', 'Height', 'Weight', 'Blood Type', 'GSIG ID', 'PagIbig ID', 
               'PhilHealth ID', 'SSS ID', 'TIN ID', 'Citizenship', 'Address', 'Telephone #',
               'Cellphone #', 'Created At', 'Updated At', 'Is Active']
    
    # Define a fill color for the header (e.g., light blue)
    header_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
    
    # Add headers to the first row and make them bold with a background color and center alignment
    for col_num, header in enumerate(headers, 1):
        cell = sheet.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)  # Make the header bold
        cell.fill = header_fill  # Apply background color
        cell.alignment = Alignment(horizontal="center", vertical="center")  # Center-align the text

    # Query all grade levels
    staffs = Staffs.objects.all()

    # Add data rows starting from row 2 with center alignment
    for row_num, staff in enumerate(staffs, 2):
        sheet.cell(row=row_num, column=1).value = staff.admin.last_name
        sheet.cell(row=row_num, column=2).value = staff.admin.first_name
        sheet.cell(row=row_num, column=3).value = staff.middle_name
        sheet.cell(row=row_num, column=4).value = staff.admin.email
        sheet.cell(row=row_num, column=5).value = staff.age
        sheet.cell(row=row_num, column=6).value = staff.dob
        sheet.cell(row=row_num, column=7).value = staff.pob
        sheet.cell(row=row_num, column=8).value = staff.sex
        sheet.cell(row=row_num, column=9).value = staff.civil_status
        sheet.cell(row=row_num, column=10).value = staff.height
        sheet.cell(row=row_num, column=11).value = staff.weight
        sheet.cell(row=row_num, column=12).value = staff.blood_type
        sheet.cell(row=row_num, column=13).value = staff.gsis_id
        sheet.cell(row=row_num, column=14).value = staff.pagibig_id
        sheet.cell(row=row_num, column=15).value = staff.philhealth_id
        sheet.cell(row=row_num, column=16).value = staff.sss_id
        sheet.cell(row=row_num, column=17).value = staff.tin_id
        sheet.cell(row=row_num, column=18).value = staff.citizenship
        sheet.cell(row=row_num, column=19).value = staff.permanent_address
        sheet.cell(row=row_num, column=20).value = staff.telephone_no
        sheet.cell(row=row_num, column=21).value = staff.cellphone_no
        sheet.cell(row=row_num, column=22).value = staff.created_at.strftime('%Y-%m-%d %H:%M:%S')
        sheet.cell(row=row_num, column=23).value = staff.updated_at.strftime('%Y-%m-%d %H:%M:%S')
        sheet.cell(row=row_num, column=24).value = staff.admin.is_active

        # Apply center alignment to all cells in the row
        for col_num in range(1, 25):
            cell = sheet.cell(row=row_num, column=col_num)
            cell.alignment = Alignment(horizontal="center", vertical="center")

    # Adjust column widths to fit the content
    for col in sheet.columns:
        max_length = 0
        column = col[0].column_letter  # Get the column letter (A, B, C, etc.)
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = max_length + 2  # Add some padding to the width
        sheet.column_dimensions[column].width = adjusted_width

    # Prepare the response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=StaffRecord.xlsx'

    # Save the workbook to the response
    workbook.save(response)

    return response

def export_students_to_excel(request):
    # Create an in-memory workbook
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = 'Students'
    
    # Define the headers
    headers = ['Last Name', 'First Name', 'Middle Name', 'Email', 'Grade Level', 'Age', 'Date of Birth', 'Place of Birth', 
               'Sex', 'Nationality', 'Religion', 'Rank In Family', 'Telephone #',
               'Mobile #', 'Start Semester', 'End Semester', 'Created At', 'Updated At', 'Is Active']
    
    # Define a fill color for the header (e.g., light blue)
    header_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
    
    # Add headers to the first row and make them bold with a background color and center alignment
    for col_num, header in enumerate(headers, 1):
        cell = sheet.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)  # Make the header bold
        cell.fill = header_fill  # Apply background color
        cell.alignment = Alignment(horizontal="center", vertical="center")  # Center-align the text

    # Query all grade levels
    students = Students.objects.all()

    # Add data rows starting from row 2 with center alignment
    for row_num, student in enumerate(students, 2):
        sheet.cell(row=row_num, column=1).value = student.admin.last_name
        sheet.cell(row=row_num, column=2).value = student.admin.first_name
        sheet.cell(row=row_num, column=3).value = student.middle_name
        sheet.cell(row=row_num, column=4).value = student.admin.email
        sheet.cell(row=row_num, column=5).value = student.GradeLevel_id.GradeLevel_name
        sheet.cell(row=row_num, column=6).value = student.age
        sheet.cell(row=row_num, column=7).value = student.dob
        sheet.cell(row=row_num, column=8).value = student.sex
        sheet.cell(row=row_num, column=9).value = student.pob
        sheet.cell(row=row_num, column=10).value = student.nationality
        sheet.cell(row=row_num, column=11).value = student.religion
        sheet.cell(row=row_num, column=12).value = student.rank_in_family
        sheet.cell(row=row_num, column=13).value = student.telephone_nos
        sheet.cell(row=row_num, column=14).value = student.mobile_phone_nos
        sheet.cell(row=row_num, column=15).value = student.session_year_id.session_start_year
        sheet.cell(row=row_num, column=16).value = student.session_year_id.session_end_year
        sheet.cell(row=row_num, column=17).value = student.created_at.strftime('%Y-%m-%d %H:%M:%S')
        sheet.cell(row=row_num, column=18).value = student.updated_at.strftime('%Y-%m-%d %H:%M:%S')
        sheet.cell(row=row_num, column=19).value = student.admin.is_active

        # Apply center alignment to all cells in the row
        for col_num in range(1, 20):
            cell = sheet.cell(row=row_num, column=col_num)
            cell.alignment = Alignment(horizontal="center", vertical="center")

    # Adjust column widths to fit the content
    for col in sheet.columns:
        max_length = 0
        column = col[0].column_letter  # Get the column letter (A, B, C, etc.)
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = max_length + 2  # Add some padding to the width
        sheet.column_dimensions[column].width = adjusted_width

    # Prepare the response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=StudentRecord.xlsx'

    # Save the workbook to the response
    workbook.save(response)

    return response


def manage_gradelevel(request):
    gradelevels = GradeLevel.objects.all()
    context = {
        "gradelevels": gradelevels
    }
    return render(request, 'hod_template/Manage_Template/manage_gradelevel_template.html', context)


def edit_gradelevel(request, GradeLevel_id):
    gradelevel = GradeLevel.objects.get(id=GradeLevel_id)
    context = {
        "gradelevel": gradelevel,
        "id": GradeLevel_id
    }
    return render(request, 'hod_template/Edit_Template/edit_gradelevel_template.html', context)


def edit_gradelevel_save(request):
    if request.method != "POST":
        HttpResponse("Invalid Method")
    else:
        GradeLevel_id = request.POST.get('GradeLevel_id')
        GradeLevel_name = request.POST.get('gradelevel')

        try:
            gradelevel = GradeLevel.objects.get(id=GradeLevel_id)
            gradelevel.GradeLevel_name = GradeLevel_name
            gradelevel.save()

            messages.success(request, "GradeLevel Updated Successfully.")
            return redirect('/edit_gradelevel/'+GradeLevel_id)

        except:
            messages.error(request, "Failed to Update GradeLevel.")
            return redirect('/edit_gradelevel/'+GradeLevel_id)

def delete_gradelevel(request, GradeLevel_id):
    gradelevel = GradeLevel.objects.get(id=GradeLevel_id)
    try:
        gradelevel.delete()
        messages.success(request, "GradeLevel Deleted Successfully.")
        return redirect('manage_gradelevel')
    except:
        messages.error(request, "Failed to Delete GradeLevel.")
        return redirect('manage_gradelevel')

def manage_session(request):
    session_years = SessionYearModel.objects.filter(is_archived=False)
    context = {
        "session_years": session_years
    }
    return render(request, "hod_template/Manage_Template/manage_session_template.html", context)

def add_session(request):
    # Get the latest session year if it exists
    latest_session = SessionYearModel.objects.order_by('-session_end_year').first()

    if latest_session:
        # Use the end year of the last session as the start year of the new session
        last_end_year = latest_session.session_end_year.year
        # Set the new session years based on the last session
        session_start_year = f"{last_end_year}-06-01"  # June 1 of the last session's end year
        session_end_year = f"{last_end_year + 1}-03-31"  # March 31 of the next year
    else:
        # If no session exists, set default values, e.g., for the current year
        current_year = datetime.now().year
        session_start_year = f"{current_year}-06-01"  # June 1 of the current year
        session_end_year = f"{current_year + 1}-03-31"  # March 31 of the next year

    print(f"Session Start Year: {session_start_year}, Session End Year: {session_end_year}")  # Debug print

    return render(request, "hod_template/Add_Template/add_session_template.html", {
        'session_start_year': session_start_year,
        'session_end_year': session_end_year
    })

def add_session_save(request):
    if request.method != "POST":
        messages.error(request, "Invalid Method")
        return redirect('add_session')
    else:
        session_start_year = request.POST.get('session_start_year')
        session_end_year = request.POST.get('session_end_year')
        session_limit = request.POST.get('session_limit')
        session_status = "Active"

        try:
            sessionyear = SessionYearModel(
                session_start_year = session_start_year, 
                session_end_year = session_end_year,
                session_limit = session_limit,
                session_status = session_status
                )
            sessionyear.save()
            messages.success(request, "Session Year added Successfully!")
            return redirect("add_session")
        except:
            messages.error(request, "Failed to Add Session Year")
            return redirect("add_session")

def edit_session(request, session_id):
    session_year = SessionYearModel.objects.get(id=session_id)
    context = {
        "session_year": session_year
    }
    return render(request, "hod_template/Edit_Template/edit_session_template.html", context)

def edit_session_save(request):
    if request.method != "POST":
        messages.error(request, "Invalid Method!")
        return redirect('manage_session')
    else:
        session_id = request.POST.get('session_id')
        session_start_year = request.POST.get('session_start_year')
        session_end_year = request.POST.get('session_end_year')
        session_limit = request.POST.get('session_limit')
        session_status = request.POST.get('session_status')

        try:
            session_year = SessionYearModel.objects.get(id=session_id)
            session_year.session_start_year = session_start_year
            session_year.session_end_year = session_end_year
            session_year.session_limit = session_limit
            session_year.session_status = session_status
            session_year.save()

            messages.success(request, "Session Year Updated Successfully.")
            return redirect('manage_session')
        except Exception as e:
            messages.error(request, f"Failed to Update Session Year: {e}")
            return redirect('manage_session')

def archived_sessions(request):
    # Filter session years with status 'Archived'
    session_years = SessionYearModel.objects.filter(is_archived=True)
    context = {
        "session_years": session_years
    }
    return render(request, 'hod_template/Archive_Template/archived_session_template.html', context)

def archive_session(request, session_id):
    session_year = get_object_or_404(SessionYearModel, id=session_id)
    session_year.is_archived = True
    session_year.session_status = "Inactive"
    session_year.save()
    messages.success(request, "Session archived successfully!")
    return redirect('manage_session')

def unarchive_session(request, session_id):
    session_year = get_object_or_404(SessionYearModel, id=session_id)
    session_year.is_archived = False
    session_year.session_status = "Active"
    session_year.save()
    messages.success(request, "Session unarchived successfully!")
    return redirect('manage_session')

def delete_session(request, session_id):
    session = SessionYearModel.objects.get(id=session_id)
    try:
        session.delete()
        messages.success(request, "Session Deleted Successfully.")
        return redirect('manage_session')
    except:
        messages.error(request, "Failed to Delete Session.")
        return redirect('manage_session')

def add_student(request):
    gradelevels = GradeLevel.objects.all()
    sessions = SessionYearModel.objects.all()
    year = datetime.now().year

    # Debug print to check current year
    print(f"Current Year: {year}")

    # Fetch the last student number starting with the current year
    last_student = Students.objects.filter(student_number__startswith=f"{year}-").order_by('-student_number').first()

    if last_student:
        # Debug print to check the last student number fetched
        print(f"Last Student Number: {last_student.student_number}")
        
        # Extract the last four digits after the dash, convert to an integer, and increment by 1
        last_number = int(last_student.student_number.split('-')[1])
        new_number = f"{year}-{str(last_number + 1).zfill(4)}"
    else:
        # Start with 0001 if no student exists for the current year
        new_number = f"{year}-0001"

    # Debug print to check the new student number
    print(f"New Student Number: {new_number}")

    context = {
        "gradelevels": gradelevels,
        "sessions": sessions,
        "student_number": new_number
    }
    return render(request, 'hod_template/Add_Template/add_student2_template.html', context)

def add_student_save(request):
    if request.method != "POST":
        messages.error(request, "Method Not Allowed!")
        return redirect('add_student')

    try:
        # Extract data from request.POST
        first_name = request.POST.get('first_name')
        last_name = request.POST.get('last_name')
        username = request.POST.get('username')
        email = request.POST.get('email')
        
        middlename = request.POST.get('middle_name')
        suffix = request.POST.get("suffix")
        nickname = request.POST.get("nickname")
        address = request.POST.get('address')
        sex = request.POST.get('sex')
        age = request.POST.get('age')
        dob = request.POST.get('dob')
        pob = request.POST.get('pob')
        nationality = request.POST.get('nationality')
        religion = request.POST.get('religion')
        rank_in_family = request.POST.get('rank_in_family')
        telephone_nos = request.POST.get('telephone_nos')
        mobile_phone_nos = request.POST.get('mobile_phone_nos')
        is_covid_vaccinated = request.POST.get('is_covid_vaccinated')
        date_of_vaccination = request.POST.get('date_of_vaccination')

        # Handle file upload for profile picture
        profile_pic_url = None
        if 'profile_pic' in request.FILES:
            profile_pic = request.FILES['profile_pic']
            upload_result = cloudinary.uploader.upload(profile_pic, folder="media/")
            profile_pic_url = upload_result.get('url')

        # Generate a predefined password based on the student's details
        current_year = now().year
        predefined_password = f"{first_name.lower()}.{last_name.lower()}.{current_year}"

        # Generate student_number
        last_student = Students.objects.filter(student_number__startswith=str(current_year)).order_by('-id').first()
        if last_student:
            last_number = int(last_student.student_number.split('-')[1])
            new_number = last_number + 1
        else:
            new_number = 1
        student_number = f"{current_year}-{new_number:04d}"

        # Create Student User
        user = CustomUser.objects.create_user(
            username=username, 
            password=(predefined_password),  # Use hashed password
            email=email,
            first_name=first_name, 
            last_name=last_name, 
            user_type=3
        )
        user.students.student_number = student_number  
        user.students.middle_name = middlename
        user.students.suffix = suffix
        user.students.nickname = nickname
        user.students.sex = sex
        user.students.address = address
        user.students.profile_pic = profile_pic_url
        user.students.age = age
        user.students.dob = dob
        user.students.pob = pob
        user.students.nationality = nationality
        user.students.religion = religion
        user.students.rank_in_family = rank_in_family
        user.students.telephone_nos = telephone_nos
        user.students.mobile_phone_nos = mobile_phone_nos
        user.students.is_covid_vaccinated = is_covid_vaccinated
        user.students.date_of_vaccination = date_of_vaccination
        user.students.student_status = "Pending"
        user.save()

        # Save Parent/Guardian Information
        ParentGuardian.objects.create(
            students_id=user.students,
            father_name=request.POST.get('father_name'),
            father_occupation=request.POST.get('father_occupation'),
            mother_name=request.POST.get('mother_name'),
            mother_occupation=request.POST.get('mother_occupation'),
            guardian_name=request.POST.get('guardian_name'),
            guardian_occupation=request.POST.get('guardian_occupation')
        )

        # Save Previous School Information
        PreviousSchool.objects.create(
            students_id=user.students,
            previous_school_name=request.POST.get('previous_school_name'),
            previous_school_address=request.POST.get('previous_school_address'),
            previous_grade_level=request.POST.get('previous_grade_level'),
            previous_school_year_attended=request.POST.get('previous_school_year_attended'),
            previous_teacher_name=request.POST.get('previous_teacher_name')
        )

        # Save Emergency Contact Information
        EmergencyContact.objects.create(
            students_id=user.students,
            emergency_contact_name=request.POST.get('emergency_contact_name'),
            emergency_contact_relationship=request.POST.get('emergency_contact_relationship'),
            emergency_contact_address=request.POST.get('emergency_contact_address'),
            emergency_contact_phone=request.POST.get('emergency_contact_phone'),
            emergency_enrolling_teacher=request.POST.get('emergency_enrolling_teacher'),
            emergency_date=request.POST.get('emergency_date'),
            emergency_referred_by=request.POST.get('emergency_referred_by')
        )

        messages.success(request, "Student Added Successfully! Password: " + predefined_password)
        return redirect('add_student')

    except Exception as e:
        messages.error(request, f"Failed to Add Student! Error: {str(e)}")
        return redirect('add_student')

def manage_student(request):
    students = Students.objects.all()

    all_users = []

    # Retrieve and structure data for Students
    for student in students:
        # Retrieve the enrollment record for the student
        enrollment = Enrollment.objects.filter(student_id=student).first()

        # Retrieve associated user data
        student_data = {
            'user': student.admin,
            'user_id': student.admin.id,
            'first_name': student.admin.first_name,  
            'last_name': student.admin.last_name,    
            'username': student.admin.username,      
            'email': student.admin.email,            
            'last_login': student.admin.last_login,  
            'date_joined': student.admin.date_joined, 
            'user_type': student.admin.user_type,
            'student_number': student.student_number,
            'middle_name': student.middle_name,
            'suffix': student.suffix,
            'nickname': student.nickname,
            'dob': student.dob,
            'age': student.age,
            'pob': student.pob,
            'sex': student.sex,
            'address': student.address,
            'religion': student.religion,
            'rank_in_family': student.rank_in_family,
            'telephone_nos': student.telephone_nos,
            'mobile_phone_nos': student.mobile_phone_nos,
            'is_covid_vaccinated': student.is_covid_vaccinated,
            'date_of_vaccination': student.date_of_vaccination,
            'student_status': student.student_status,
            'parent_guardian': ParentGuardian.objects.filter(students_id=student).first(),
            'previous_school': PreviousSchool.objects.filter(students_id=student).first(),
            'emergency_contact': EmergencyContact.objects.filter(students_id=student).first(),
            'grade_level': enrollment.GradeLevel_id.GradeLevel_name if enrollment else None,
            'session_year': (
                f"{enrollment.session_year_id.session_start_year.year}-{enrollment.session_year_id.session_end_year.year}"
                if enrollment else None
            ),
            'enrollment': enrollment,
        }
        all_users.append(student_data)

    context = {
        'all_users': all_users
    }
    return render(request, 'hod_template/Manage_Template/manage_student_template.html', context)


def toggle_student_activation(request, user_id):
    user = get_user_model().objects.get(id=user_id)

    if user.is_active:
        user.is_active = False
        messages.success(request, f"User {user.username} has been deactivated.")
    else:
        user.is_active = True
        messages.success(request, f"User {user.username} has been reactivated.")
    
    user.save()

    return redirect('manage_student')

def add_enrollment(request):
    # Get all grade levels that are not already associated with an Enrollment_voucher
    enrolled_grade_levels = Enrollment_voucher.objects.values('GradeLevel_id_id')
    gradelevels = GradeLevel.objects.exclude(id__in=Subquery(enrolled_grade_levels))
    context = {
        "gradelevels": gradelevels,
    }
    return render(request, 'hod_template/Add_Template/add_enrollment_template.html', context)

def add_enrollment_save(request):
    if request.method != "POST":
        messages.error(request, "Invalid Method!")
        return redirect('add_enrollment')

    # Retrieve and convert values from the form
    def safe_decimal(value, field_name):
        """Helper function to safely convert to Decimal with error logging."""
        try:
            return Decimal(value.strip() or '0.00')
        except InvalidOperation:
            messages.error(request, f"Invalid input for {field_name}. Please check your values.")
            raise

    try:
        # Convert and log values
        GradeLevel_id = request.POST.get('GradeLevel_id')
        registration_fee = safe_decimal(request.POST.get('registration_fee', '0.00'), 'registration fee')
        misc_fee = safe_decimal(request.POST.get('misc_fee', '0.00'), 'misc fee')
        tuition_fee = safe_decimal(request.POST.get('tuition_fee', '0.00'), 'tuition fee')
        total_fee = safe_decimal(request.POST.get('total_fee', '0.00'), 'total fee')
        
        # Create and save the enrollment instance
        enrollment = Enrollment_voucher(
            GradeLevel_id = GradeLevel.objects.get(id=GradeLevel_id),
            registration_fee=registration_fee,
            misc_fee=misc_fee,
            tuition_fee=tuition_fee,
            total_fee=total_fee,
        )
        
        enrollment.save()
        messages.success(request, "Enrollment Added Successfully!")
    except Exception as e:
        messages.error(request, f"Failed to Add Enrollment! Error: {str(e)}")
    return redirect('add_enrollment')

def manage_enrollment(request):
    # Filter students with "Pending" enrollment status
    students = Students.objects.filter(
        enrollment__enrollment_status="Pending"
    ).distinct()  # Ensure we only get students with a "Pending" enrollment status

    # Create a dictionary of vouchers by GradeLevel_id
    enrollment_vouchers = {voucher.GradeLevel_id.id: voucher for voucher in Enrollment_voucher.objects.all()}

    # Annotate each student with the relevant voucher and grade level based on GradeLevel_id
    for student in students:
        # Get the corresponding enrollment for the student to check the GradeLevel_id
        enrollment = Enrollment.objects.filter(student_id=student.id, enrollment_status="Pending").first()
        if enrollment:
            student.voucher = enrollment_vouchers.get(enrollment.GradeLevel_id.id)
            student.grade_level_name = enrollment.GradeLevel_id.GradeLevel_name  # Add grade level to student
            student.enrollment_status = enrollment.enrollment_status  # Add enrollment status to student

    school = School_info.objects.first()
    context = {
        "students": students,
        "school": school,
        'now': timezone.now(),
    }
    return render(request, 'hod_template/Manage_Template/manage_enrollment_template.html', context)

def update_student_status(request):
    # Handle GET request: Show confirmation or handle any required logic
    if request.method == 'GET':
        student_id = request.GET.get('student_id')
        
        if student_id:
            try:
                # Find the student by ID
                student = Students.objects.get(id=student_id)
                
                # Find the enrollment record for the student
                enrollment = Enrollment.objects.filter(student_id=student).first()
                
                if enrollment:
                    # Update the enrollment status
                    enrollment.enrollment_status = 'Enrolled'  # Update enrollment status
                    enrollment.save()

                    # Add a success message to be shown after redirection
                    messages.success(request, 'Enrollment status updated successfully.')
                else:
                    messages.error(request, 'Enrollment record not found for this student.')

                # Redirect to a page, like managing the enrollment
                return redirect('manage_enrollment')
            
            except Students.DoesNotExist:
                # If the student doesn't exist, add an error message
                messages.error(request, 'Student not found.')
                return redirect('manage_enrollment')
        
        # If no student_id is passed, return an error
        messages.error(request, 'Invalid student ID.')
        return redirect('manage_enrollment')

    # If the method is POST, handle it as your original code
    if request.method == 'POST':
        student_id = request.POST.get('student_id')
        try:
            student = Students.objects.get(id=student_id)
            # Find the enrollment record for the student
            enrollment = Enrollment.objects.filter(student_id=student).first()
            
            if enrollment:
                enrollment.enrollment_status = "Enrolled"  # Update enrollment status
                enrollment.save()
                messages.success(request, 'Enrollment status updated successfully.')
            else:
                messages.error(request, 'Enrollment record not found for this student.')
        except Students.DoesNotExist:
            messages.error(request, 'Student not found.')

        return redirect('manage_enrollment')

    # For any other method (if any), return a response
    return HttpResponse("Invalid request method.", status=405)

def manage_enrollment_voucher(request):
    enrollment_vouchers = Enrollment_voucher.objects.all()
    context = {
        "enrollment_vouchers": enrollment_vouchers,
    }
    return render(request, 'hod_template/Manage_Template/manage_enrollment_voucher_template.html', context)

def edit_enrollment(request, enrollment_id):
    enrollment = Enrollment_voucher.objects.get(id=enrollment_id)
    gradelevels = GradeLevel.objects.all()

    context = {
        "enrollment": enrollment,
        "gradelevels": gradelevels,
    }

    return render(request, 'hod_template/Edit_Template/edit_enrollment_template.html', context)

def edit_enrollment_save(request, enrollment_id):
    if request.method != "POST":
        return HttpResponse("Invalid Method.")
    else:
        # Retrieve other POST data
        registration_fee = request.POST.get('registration_fee')
        misc_fee = request.POST.get('misc_fee')
        tuition_fee = request.POST.get('tuition_fee')
        total_fee = request.POST.get('total_fee')

        try:
            enrollment = Enrollment_voucher.objects.get(id=enrollment_id)
            enrollment.registration_fee = registration_fee
            enrollment.misc_fee = misc_fee
            enrollment.tuition_fee = tuition_fee
            enrollment.total_fee = total_fee

            enrollment.save()

            messages.success(request, "Enrollment Voucher Updated Successfully.")
            # Redirect to the manage_enrollment_voucher page after success
            return HttpResponseRedirect(reverse("manage_enrollment_voucher"))
        except Enrollment_voucher.DoesNotExist:
            messages.error(request, "Enrollment record not found.")
        except Exception as e:
            messages.error(request, f"Failed to Update Enrollment: {e}")
        
        # If there was an error, redirect back to the edit enrollment page
        return HttpResponseRedirect(reverse("edit_enrollment", kwargs={"enrollment_id": enrollment_id}))


def edit_student(request, student_id):
    student = get_object_or_404(Students, id=student_id)
    sessions = SessionYearModel.objects.all()
    gradelevels = GradeLevel.objects.all()
    parentguardian = ParentGuardian.objects.filter(students_id=student).first()
    previousschool = PreviousSchool.objects.filter(students_id=student).first()
    emergencycontact = EmergencyContact.objects.filter(students_id=student).first()

    context = {
        "student": student,
        "rank_in_family": student.rank_in_family,
        "sessions": sessions,
        "gradelevels": gradelevels,
        "parentguardian": parentguardian, 
        "previousschool": previousschool,
        "emergencycontact": emergencycontact,
    }

    return render(request, 'hod_template/Edit_Template/edit_student_template.html', context)

def edit_student_save(request, student_id):
    if request.method != "POST":
        messages.error(request, "Method Not Allowed!")
        return redirect('edit_student', student_id=student_id)

    try:
        # Fetch the student based on the provided student_id
        student = Students.objects.get(id=student_id)
        user = student.admin  # The associated CustomUser instance (admin)

        # Extract data from request.POST
        first_name = request.POST.get('first_name')
        last_name = request.POST.get('last_name')
        username = request.POST.get('username')
        middlename = request.POST.get('middle_name')
        suffix = request.POST.get("suffix")
        nickname = request.POST.get("nickname")
        email = request.POST.get('email')
        address = request.POST.get('address')
        session_year_id = request.POST.get('session_year_id')
        grade_level_id = request.POST.get('GradeLevel_id')
        sex = request.POST.get('sex')
        age = request.POST.get('age')
        dob = request.POST.get('dob')
        pob = request.POST.get('pob')
        nationality = request.POST.get('nationality')
        religion = request.POST.get('religion')
        rank_in_family = request.POST.get('rank_in_family')
        telephone_nos = request.POST.get('telephone_nos')
        mobile_phone_nos = request.POST.get('mobile_phone_nos')
        is_covid_vaccinated = request.POST.get('is_covid_vaccinated') == 'on'
        date_of_vaccination = request.POST.get('date_of_vaccination')

        # Handle file upload for profile picture
        profile_pic_url = None
        if 'profile_pic' in request.FILES:
            profile_pic = request.FILES['profile_pic']
            upload_result = cloudinary.uploader.upload(profile_pic, folder="media/")
            profile_pic_url = upload_result.get('url')

        # Update user details
        user.username = username
        user.email = email
        user.first_name = first_name
        user.last_name = last_name
        user.save()  # Save the CustomUser instance

        # Update student details
        student.middle_name = middlename
        student.suffix = suffix
        student.nickname = nickname
        student.address = address
        student.sex = sex
        student.age = age
        student.dob = dob
        student.pob = pob
        student.nationality = nationality
        student.religion = religion
        student.rank_in_family = rank_in_family
        student.telephone_nos = telephone_nos
        student.mobile_phone_nos = mobile_phone_nos
        student.is_covid_vaccinated = is_covid_vaccinated
        student.date_of_vaccination = date_of_vaccination

        if profile_pic_url:
            student.profile_pic = profile_pic_url
        
        # Assign GradeLevel and Session Year
        gradelevel_obj = GradeLevel.objects.get(id=grade_level_id)
        session_year_obj = SessionYearModel.objects.get(id=session_year_id)
        student.GradeLevel_id = gradelevel_obj
        student.session_year_id = session_year_obj

        # Save the updated student record
        student.save()

        # Update Parent/Guardian Information (if it already exists, fetch and update it)
        parent_guardian, created = ParentGuardian.objects.get_or_create(students_id=student)
        parent_guardian.father_name = request.POST.get('father_name')
        parent_guardian.father_occupation = request.POST.get('father_occupation')
        parent_guardian.mother_name = request.POST.get('mother_name')
        parent_guardian.mother_occupation = request.POST.get('mother_occupation')
        parent_guardian.guardian_name = request.POST.get('guardian_name')
        parent_guardian.guardian_occupation = request.POST.get('guardian_occupation')
        parent_guardian.save()

        # Update Previous School Information (fetch and update if it exists)
        previous_school, created = PreviousSchool.objects.get_or_create(students_id=student)
        previous_school.previous_school_name = request.POST.get('previous_school_name')
        previous_school.previous_school_address = request.POST.get('previous_school_address')
        previous_school.previous_grade_level = request.POST.get('previous_grade_level')
        previous_school.previous_school_year_attended = request.POST.get('previous_school_year_attended')
        previous_school.previous_teacher_name = request.POST.get('previous_teacher_name')
        previous_school.save()

        # Update Emergency Contact Information (fetch and update if it exists)
        emergency_contact, created = EmergencyContact.objects.get_or_create(students_id=student)
        emergency_contact.emergency_contact_name = request.POST.get('emergency_contact_name')
        emergency_contact.emergency_contact_relationship = request.POST.get('emergency_contact_relationship')
        emergency_contact.emergency_contact_address = request.POST.get('emergency_contact_address')
        emergency_contact.emergency_contact_phone = request.POST.get('emergency_contact_phone')
        emergency_contact.emergency_enrolling_teacher = request.POST.get('emergency_enrolling_teacher')
        emergency_contact.emergency_date = request.POST.get('emergency_date')
        emergency_contact.emergency_referred_by = request.POST.get('emergency_referred_by')
        emergency_contact.save()

        # Success message
        messages.success(request, "Student Information Updated Successfully!")
        return redirect('manage_student')

    except Students.DoesNotExist:
        messages.error(request, "Student Not Found!")
        return redirect('edit_student', student_id=student_id)

    except Exception as e:
        messages.error(request, f"Failed to Update Student! Error: {str(e)}")
        return redirect('edit_student', student_id=student_id)

def deactivate_student(request, student_id):
    student = Students.objects.get(admin=student_id)
    try:
        # Set the is_active field of the associated CustomUser to False
        student.admin.is_active = False
        student.admin.save()
        
        messages.success(request, "Student Deactivated Successfully.")
        return redirect('manage_student')
    except Exception as e:
        print(e)
        messages.error(request, "Failed to Deactivate Student.")
        return redirect('manage_student')

def manage_section(request):
    gradelevels = GradeLevel.objects.all()
    sections = Section.objects.all()

    context = {
        "sections": sections,
        "gradelevels":gradelevels,
    }
    return render(request, "hod_template/Manage_Template/manage_section_template.html", context)

def add_section(request):
    gradelevels = GradeLevel.objects.all()
    context = {
        "gradelevels":gradelevels,
    }
    return render(request, 'hod_template/Add_Template/add_section_template.html', context)

def add_section_save(request):
    if request.method != "POST":
        messages.error(request, "Method Not Allowed!")
        return redirect('add_section')
    else:
        section_name = request.POST.get('section_name')
        section_limit = request.POST.get('section_limit')
        section_soft_limit = request.POST.get('section_soft_limit')
        GradeLevel_id = request.POST.get('gradelevel')
        gradelevel = GradeLevel.objects.get(id=GradeLevel_id)
        
        try:
            section = Section(
                section_name = section_name,
                section_soft_limit = section_soft_limit,
                section_limit = section_limit,
                GradeLevel_id=gradelevel)
            section.save()
            messages.success(request, "Section Added Successfully!")
            return redirect('add_section')
        except:
            messages.error(request, "Failed to Add Section!")
            return redirect('add_section')

def edit_section(request, section_id):

    section = Section.objects.get(id=section_id)
    gradelevel = GradeLevel.objects.all()

    context = {
        "section": section,
        "gradelevel":gradelevel,
    }

    return render(request, 'hod_template/Edit_Template/edit_section_template.html', context)

def edit_section_save(request):
    if request.method != "POST":
        HttpResponse("Invalid Method.")
    else:
        section_id = request.POST.get('section_id')
        section_name = request.POST.get('section_name')
        section_limit = request.POST.get('section_limit')
        section_soft_limit = request.POST.get('section_soft_limit')
        GradeLevel_id = request.POST.get('gradelevel')
        
        try:
            section = Section.objects.get(id=section_id)
            section.section_name = section_name

            gradelevel = GradeLevel.objects.get(id=GradeLevel_id)
            section.GradeLevel_id = gradelevel

            section_limit = section_limit
            section_soft_limit = section_soft_limit

            section.save()

            messages.success(request, "Section Updated Successfully.")
            return HttpResponseRedirect(reverse("edit_section", kwargs={"section_id":section_id}))
        except:
            messages.error(request, "Failed to Update Section.")
            return HttpResponseRedirect(reverse("edit_section", kwargs={"section_id":section_id}))
            
def add_subject(request):
    curriculums = Curriculums.objects.all()
    gradelevels = GradeLevel.objects.all()
    context = {
        "curriculums": curriculums,
        "gradelevels": gradelevels
        
    }
    return render(request, 'hod_template/Add_Template/add_subject_template.html', context)

def add_subject_save(request):
    if request.method != "POST":
        messages.error(request, "Method Not Allowed!")
        return redirect('add_subject')
    else:
        subject_name = request.POST.get('subject_name')
        subject_description = request.POST.get('subject_description')
        subject_code = request.POST.get('subject_code')
        subject_hours = request.POST.get('subject_hours')

        curriculum_id = request.POST.get('curriculum_id')
        curriculum_id = Curriculums.objects.get(id=curriculum_id)

        GradeLevel_id = request.POST.get('GradeLevel_id')
        GradeLevel_id = GradeLevel.objects.get(id=GradeLevel_id)
        
        try:
            subject = Subjects(subject_name=subject_name,
                               subject_description=subject_description,
                               subject_code=subject_code,
                               subject_hours=subject_hours,
                               GradeLevel_id=GradeLevel_id, 
                               curriculum_id = curriculum_id)
            subject.save()
            messages.success(request, "Subject Added Successfully!")
            return redirect('add_subject')
        except:
            messages.error(request, "Failed to Add Subject!")
            return redirect('add_subject')

def manage_subject(request):
    subjects = Subjects.objects.all()
    context = {
        "subjects": subjects
    }
    return render(request, 'hod_template/Manage_Template/manage_subject_template.html', context)

def manage_classroom(request):
    classroom = Classroom.objects.all()
    context = {
        "classrooms": classroom
    }
    return render(request, 'hod_template/Manage_Template/manage_classroom_template.html', context)

def fetch_classroom(request):
    api_url = "https://inventoryapp1-o2l3.onrender.com/classroom/"

    try:
        response = requests.get(api_url)
        response.raise_for_status()  # Raise an exception for HTTP errors

        data = response.json()
        classrooms = data.get("classrooms", [])

        # Step 1: Fetch all IDs from the external API
        external_ids = [classroom["id"] for classroom in classrooms]

        # Step 2: Update or create records for all received data
        for classroom in classrooms:
            Classroom.objects.update_or_create(
                id=classroom["id"],
                defaults={
                    "classroom_name": classroom["classroom_name"],
                    "capacity": classroom["capacity"],  # Handle new 'status' field
                },
            )

        # Step 3: Identify and delete records that are no longer in the external API
        Classroom.objects.exclude(id__in=external_ids).delete()

        # Add a success message
        messages.success(request, "Classrooms synchronized successfully.")
        return redirect("manage_classroom")  # Redirect to the same page

    except requests.RequestException as e:
        messages.error(request, f"Request failed: {str(e)}")
        return redirect("manage_classroom")

    except Exception as e:
        messages.error(request, f"An unexpected error occurred: {str(e)}")
        return redirect("manage_classroom")

# class ManageSubjectsAPIView(APIView):
#     permission_classes = [IsAuthenticated]  # Only authenticated users

#     def post(self, request):
#         refresh_token = request.data.get("refresh_token")
#         if not refresh_token:
#             return Response({"error": "Refresh token is required"}, status=400)

#         try:
#             token = RefreshToken(refresh_token)

#             if token.check_blacklist():
#                 return Response({"error": "Token is blacklisted"}, status=403)
            
#             user_id = token.get("user_id")

#             user = CustomUser.objects.get(id=user_id)

#             if user.user_type != "1":
#                 return Response({"error": "You do not have permission to view this resource."}, status=403)

#             subjects = Subjects.objects.all().values(
#                 "id", 
#                 "curriculum_id__curriculum_name", 
#                 "GradeLevel_id__GradeLevel_name",
#                 "subject_name", 
#                 "subject_description", 
#                 "subject_code", 
#                 "subject_hours", 
#                 "created_at", 
#                 "updated_at"
#             )
#             return Response({"subjects": list(subjects)}, status=200)
        
#         except TokenError as e:
#             return Response({"error": f"Token error: {str(e)}"}, status=403)
#         except CustomUser.DoesNotExist:
#             return Response({"error": "User not found"}, status=404)
#         except Exception as e:
#             return Response({"error": str(e)}, status=400)

def edit_subject(request, subject_id):
    subject = Subjects.objects.get(id=subject_id)
    gradelevels = GradeLevel.objects.all()
    curriculums = Curriculums.objects.all()
    
    context = {
        "subject": subject,
        "gradelevels": gradelevels,
        "curriculums": curriculums,
        "id": subject_id
    }
    return render(request, 'hod_template/Edit_Template/edit_subject_template.html', context)

def edit_subject_save(request):
    if request.method != "POST":
        HttpResponse("Invalid Method.")
    else:
        subject_id = request.POST.get('subject_id')
        GradeLevel_id = request.POST.get('gradelevel')
        subject_name = request.POST.get('subject')
        
        subject_code = request.POST.get('subject_code')
        subject_description = request.POST.get('subject_description')
        subject_hours = request.POST.get('subject_hours')
        
        try:
            subject = Subjects.objects.get(id=subject_id)
            subject.subject_name = subject_name
            
            gradelevel = GradeLevel.objects.get(id=GradeLevel_id)
            subject.GradeLevel_id = gradelevel
            
            subject.subject_code = subject_code
            subject.subject_description = subject_description
            subject.subject_hours = subject_hours
            
            subject.save()

            messages.success(request, "Subject Updated Successfully.")
            # return redirect('/edit_subject/'+subject_id)
            return HttpResponseRedirect(reverse("edit_subject", kwargs={"subject_id":subject_id}))

        except:
            messages.error(request, "Failed to Update Subject.")
            return HttpResponseRedirect(reverse("edit_subject", kwargs={"subject_id":subject_id}))
            # return redirect('/edit_subject/'+subject_id)

def delete_subject(request, subject_id):
    subject = Subjects.objects.get(id=subject_id)
    try:
        subject.delete()
        messages.success(request, "Subject Deleted Successfully.")
        return redirect('manage_subject')
    except:
        messages.error(request, "Failed to Delete Subject.")
        return redirect('manage_subject')

def add_promotion(request):
    
    session_years = SessionYearModel.objects.filter(session_status='Active')
    all_gradelevels = GradeLevel.objects.all()

    gradelevels = GradeLevel.objects.filter(
        id__in=Enrollment.objects.filter(
            enrollment_status="Enrolled"  # Only include enrolled students by enrollment_status
        ).values_list('GradeLevel_id', flat=True)  # Use the GradeLevel_id field in Enrollment
    )

    context = {
        "session_years": session_years,
        "gradelevels": gradelevels,  
        "all_gradelevels": all_gradelevels, # All grade levels
    }
    return render(request, 'hod_template/Add_Template/add_promotion_template.html', context)

def load_promotion(request):
    session_year_id = request.GET.get('session_year_id')  # Academic Year
    gradelevel_id = request.GET.get('gradelevel_id')  # Grade Level
    section_id = request.GET.get('section_id')  # Section

    # Debugging to check the received parameters
    print("Query Parameters - session_year_id:", session_year_id)
    print("Query Parameters - gradelevel_id:", gradelevel_id)
    print("Query Parameters - section_id:", section_id)

    # Check if session_year_id and gradelevel_id are provided
    if not session_year_id or not gradelevel_id:
        return JsonResponse({'error': 'Missing required parameters: session_year_id or gradelevel_id'}, status=400)

    # Fetch sections filtered by GradeLevel if gradelevel_id is provided
    sections = Section.objects.filter(GradeLevel_id=gradelevel_id)

    # Filter students by academic year, grade level, and enrollment status
    students = Students.objects.filter(
        enrollment__session_year_id=session_year_id,  # Filter by academic year
        enrollment__GradeLevel_id=gradelevel_id,  # Filter by grade level
        enrollment__enrollment_status="Enrolled"  # Ensure students are enrolled
    )

    # If section_id is provided, filter students by section through AssignSection model
    if section_id:
        students = students.filter(
            enrollment__assignsection__section_id=section_id  # Ensure students are assigned to the selected section
        )

    print("Students Found:", students)  # Debugging

    # Prepare data for response with additional fields
    section_list = [{"id": section.id, "name": section.section_name} for section in sections]
    student_list = [
        {
            "id": student.id,
            "student_number": student.student_number,
            "first_name": student.admin.first_name,
            "middle_name": student.middle_name[0] if student.middle_name else '',  # First letter of middle name
            "last_name": student.admin.last_name,
            "student_status": student.student_status,
            "enrollment_status": Enrollment.objects.filter(student_id=student.id).first().enrollment_status  # Fetch the enrollment status
        }
        for student in students
    ]

    return JsonResponse({'sections': section_list, 'students': student_list})

def add_promotion_save(request):
    if request.method != "POST":
        messages.error(request, "Method Not Allowed!")
        return redirect('add_promotion')
    
    try:
        # Get the selected students, next session year, and next grade level
        student_ids = request.POST.getlist('student_ids')  # List of selected student IDs
        next_session_year_id = request.POST.get('next_session_year_id')
        next_gradelevel_id = request.POST.get('next_gradelevel_id')

        # Ensure students are selected
        if not student_ids:
            messages.error(request, "No students selected!")
            return redirect('add_promotion')

        # Ensure valid next session year and grade level are selected
        if not next_session_year_id or not next_gradelevel_id:
            messages.error(request, "Please select a valid next session year and grade level!")
            return redirect('add_promotion')

        # Retrieve the next session year and grade level
        next_session_year = SessionYearModel.objects.get(id=next_session_year_id)
        next_gradelevel = GradeLevel.objects.get(id=next_gradelevel_id)

        # Create new Enrollment records for the selected students
        for student_id in student_ids:
            student = Students.objects.get(id=student_id)

            # Create a new Enrollment record for the student in the next session year and grade level
            Enrollment.objects.create(
                student_id=student,
                session_year_id=next_session_year,
                GradeLevel_id=next_gradelevel,
                enrollment_status='Pending'  # Default status for new enrollment
            )

            # Update the student's status to Active if it's not already
            if student.student_status != 'Active':
                student.student_status = 'Active'
                student.save()

        messages.success(request, "Selected students have been promoted successfully!")
        return redirect('add_promotion')  # Redirect to a success page or the same form

    except Exception as e:
        messages.error(request, f"Error: {str(e)}")
        return redirect('add_promotion')

def add_admission(request):
    
    session_years = SessionYearModel.objects.filter(session_status='Active')
    gradelevels = GradeLevel.objects.all()
    # Filter students with Admission status
    students = Students.objects.filter(student_status="Admission")
    context = {
        "session_years": session_years,
        "gradelevels": gradelevels,
        "students": students,
    }
    return render(request, 'hod_template/Add_Template/add_admission_template.html', context)

def add_admission_save(request):
    if request.method != "POST":
        messages.error(request, "Method Not Allowed!")
        return redirect('add_admission')
    
    try:
        # Retrieve selected session year and grade level
        session_year_id = request.POST.get('session_year_id')
        session_year = SessionYearModel.objects.get(id=session_year_id)

        gradelevel_id = request.POST.get('GradeLevel_id')
        gradelevel = GradeLevel.objects.get(id=gradelevel_id)

        # Get the selected student IDs from the form
        student_ids = request.POST.getlist('student_ids')  # Get list of selected student IDs

        if not student_ids:
            messages.error(request, "No students selected!")
            return redirect('add_admission')

        # Save the Enrollment records for each selected student and update student status
        for student_id in student_ids:
            student = Students.objects.get(id=student_id)

            # Create an Enrollment record
            Enrollment.objects.create(
                student_id=student,
                session_year_id=session_year,
                GradeLevel_id=gradelevel,
                enrollment_status='Pending'  # Default status
            )

            # Update student status to Active
            student.student_status = 'Active'
            student.save()

        messages.success(request, "Selected students have been enrolled and status updated to Active!")
        return redirect('add_admission')  # Redirect to a success page or the same form
    
    except Exception as e:
        messages.error(request, f"Error: {str(e)}")
        return redirect('add_admission')

def add_assignsection(request):
    session_years = SessionYearModel.objects.all()

    # Filter grade levels for students with enrollment_status="Enrolled"
    gradelevels = GradeLevel.objects.filter(
        id__in=Enrollment.objects.filter(
            enrollment_status="Enrolled"  # Only include enrolled students by enrollment_status
        ).values_list('GradeLevel_id', flat=True)  # Use the GradeLevel_id field in Enrollment
    )
    
    context = {
        "session_years": session_years,
        "gradelevels": gradelevels,
    }

    return render(request, 'hod_template/Add_Template/add_assignsection_template.html', context)

def load_sections_and_students(request):
    gradelevel_id = request.GET.get('gradelevel_id')
    
    # Fetch sections and students filtered by GradeLevel
    sections = Section.objects.filter(GradeLevel_id=gradelevel_id)

    # Get all students who are assigned to sections in this grade level
    assigned_students = AssignSection.objects.filter(
        enrollment__GradeLevel_id=gradelevel_id  # Filter AssignSection by GradeLevel
    ).values_list('enrollment__student_id', flat=True)  # Access student_id through enrollment
    
    # Fetch students who are enrolled in the given grade level
    students = Students.objects.filter(
        enrollment__GradeLevel_id=gradelevel_id,  # Use Enrollment to filter by GradeLevel_id
        enrollment__enrollment_status="Enrolled"  # Only include enrolled students
    ).exclude(id__in=assigned_students)  # Exclude students already assigned to a section

    # Prepare data for response
    section_list = [{"id": section.id, "name": section.section_name} for section in sections]
    student_list = [{"id": student.id, "name": f"{student.admin.first_name} {student.admin.last_name}"} for student in students]
    
    return JsonResponse({'sections': section_list, 'students': student_list})

def add_assignsection_save(request):
    if request.method != "POST":
        messages.error(request, "Method Not Allowed!")
        return redirect('add_assignsection')
    try:
        # Get session year and grade level from the form
        session_year_id = request.POST.get('session_year_id')
        session_id = SessionYearModel.objects.get(id=session_year_id)

        gradelevel_id = request.POST.get('GradeLevel_id')
        gradelevel = GradeLevel.objects.get(id=gradelevel_id)

        section_id = request.POST.get('section_id')
        section = Section.objects.get(id=section_id)

        # Get the list of student_ids
        student_ids = request.POST.getlist('student_ids')

        if not student_ids:
            messages.error(request, "No students selected!")
            return redirect('add_assignsection')

        # Check the current count of students in this section
        current_count = AssignSection.objects.filter(section_id=section).count()
        remaining_slots = section.section_limit - current_count

        # Filter students to add based on remaining slots
        students_to_add = student_ids[:remaining_slots]  # Students that can be added within the limit
        students_not_added = student_ids[remaining_slots:]  # Students exceeding the limit

        # Proceed to save the assignment for each student within the limit
        for student_id in students_to_add:
            student = Students.objects.get(id=student_id)
            
            # Create an enrollment for the student if it doesn't exist
            enrollment, created = Enrollment.objects.get_or_create(
                student_id=student,
                session_year_id=session_id,
                GradeLevel_id=gradelevel,
                enrollment_status='Enrolled'  # Adjust status if necessary
            )
            
            # Now create the AssignSection entry linking the Enrollment with the Section
            assignsection = AssignSection(
                enrollment=enrollment,
                section_id=section
            )
            assignsection.save()

            # Prepare a success message
            success_message = f"{len(students_to_add)} students assigned successfully to section '{section.section_name}'."
            
            # If there are students that couldn't be added, prepare a warning message
            if students_not_added:
                not_added_names = [Students.objects.get(id=student_id).admin.get_full_name() for student_id in students_not_added]
                warning_message = f"However, the following students could not be assigned due to the section limit of {section.section_limit}: {', '.join(not_added_names)}."
                messages.warning(request, warning_message)
            
            messages.success(request, success_message)
            return redirect('add_assignsection')

    except Exception as e:
            messages.error(request, f"Failed to Assign Section! Error: {e}")
            return redirect('add_assignsection')

def manage_assign_section(request):
    # Fetch all session years
    session_years = SessionYearModel.objects.all()

    # Fetch all grade levels (no filtering)
    gradelevels = GradeLevel.objects.all()

    # Filter enrollments for students with student_status="Enrolled"
    enrollments = Enrollment.objects.filter(
        student_id__student_status="Enrolled"  # Correct field name is student_id
    )

    # Get all assign sections
    assignsections = AssignSection.objects.all()

    context = {
        "session_years": session_years,
        "gradelevels": gradelevels,  # Pass all grade levels
        "assignsections": assignsections,
    }

    return render(request, 'hod_template/Manage_Template/manage_assignsection_template.html', context)

def get_sections_by_grade(request, grade_level_id):
    # Try to convert grade_level_id to an integer to handle numeric grade level IDs
    try:
        grade_level_id = int(grade_level_id)
        # If it converts successfully, filter by numeric grade_level_id
        sections = Section.objects.filter(GradeLevel_id=grade_level_id)
    except ValueError:
        # If conversion fails, treat grade_level_id as a string and filter by name
        sections = Section.objects.filter(GradeLevel__name=grade_level_id)
    
    # Prepare the response
    section_data = [{'id': section.id, 'section_name': section.section_name} for section in sections]
    return JsonResponse({'sections': section_data})

def filter_assign_sections(request):
    academic_year = request.GET.get('academic_year')
    grade_level = request.GET.get('grade_level')
    section = request.GET.get('section')

    # Build filters
    filters = {}
    if academic_year:
        filters['enrollment__session_year_id'] = academic_year
    if grade_level:
        filters['enrollment__GradeLevel_id'] = grade_level
    if section:
        filters['section_id'] = section

    # Query the database with related fields
    assigned_sections = AssignSection.objects.filter(
        **filters
    ).select_related('enrollment__student_id__admin', 'enrollment__GradeLevel_id', 'section_id')

    # Prepare the JSON response
    data = {
        "assignsections": [
            {
                "id": assign.id,
                "student_name": f"{assign.enrollment.student_id.admin.first_name} {assign.enrollment.student_id.admin.last_name}",
                "grade_level_id": assign.enrollment.GradeLevel_id.id,
                "grade_level_name": assign.enrollment.GradeLevel_id.GradeLevel_name,
                "section_id": assign.section_id.id if assign.section_id else None,
                "section_name": assign.section_id.section_name if assign.section_id else "",
                "date_assigned": assign.created_at.strftime('%Y-%m-%d'),
            }
            for assign in assigned_sections
        ]
    }
    return JsonResponse(data)

def edit_assignsection(request):
    if request.method == 'POST':
        assignsection_id = request.POST.get('id')
        section_id = request.POST.get('section_id')

        try:
            assign_section = AssignSection.objects.get(id=assignsection_id)
            assign_section.section_id_id = section_id
            assign_section.save()
            messages.success(request, "Section updated successfully!")
        except AssignSection.DoesNotExist:
            messages.error(request, "Assigned section not found.")
        except Exception as e:
            messages.error(request, f"An error occurred: {e}")

        return redirect('manage_assign_section')

def add_load(request):
    # Fetch all necessary data for the context
    session_years = SessionYearModel.objects.all()
    curriculums = Curriculums.objects.all()
    
    # Fetch distinct AssignSection records with proper filters
    assignsections = AssignSection.objects.select_related(
        'enrollment__GradeLevel_id', 'section_id'
    ).distinct(
        'enrollment__GradeLevel_id', 'section_id'
    )
    
    gradelevels = GradeLevel.objects.all()
    sections = Section.objects.all()
    subjects = Subjects.objects.all()
    
    # Fetch Load records with necessary related data
    loads = Load.objects.select_related(
        'curriculum_id', 'AssignSection_id', 'subject_id', 'staff_id'
    ).all()
    
    # Fetch staff users
    staffs = CustomUser.objects.filter(user_type='2')
    # Fetching existing loads to display in the table
    loads = Load.objects.select_related('curriculum_id', 'AssignSection_id', 'subject_id', 'staff_id').all()  
    context = {
        "session_years": session_years,
        "curriculums": curriculums,
        "assignsections": assignsections,
        "gradelevels": gradelevels,
        "sections": sections,
        "subjects": subjects,
        "staffs": staffs,
        "loads": loads,
    }
    return render(request, 'hod_template/Add_Template/add_load_template.html', context)

def is_load_exists(request):
    if request.method == 'POST':
        session_year_id = request.POST.get('session_year_id')
        curriculum_id = request.POST.get('curriculum_id')
        assign_section_id = request.POST.get('assign_section_id')

        # Check if a Load record with the selected session_year, curriculum, and assign_section already exists
        exists = Load.objects.filter(
            session_year_id=session_year_id,
            curriculum_id=curriculum_id,
            AssignSection_id=assign_section_id  # Updated to match field name
        ).exists()

        # Return a JSON response indicating whether the record exists
        return JsonResponse({'exists': exists})

def get_subject_data(request):
    curriculum_id = request.GET.get('curriculum_id')
    gradelevel_id = request.GET.get('gradelevel_id')

    if curriculum_id and gradelevel_id:
        # Fetch subjects based on curriculum and grade level
        subjects = Subjects.objects.filter(curriculum_id=curriculum_id, GradeLevel_id=gradelevel_id)
        
        # Fetch all staff for the faculty dropdown (assuming user_type='2' for staff)
        staffs = CustomUser.objects.filter(user_type='2')

        # Prepare the subject data
        subject_data = [
            {"id": subject.id, "name": subject.subject_name} for subject in subjects
        ]

        # Prepare the staff data
        staff_data = [
            {"id": staff.id, "name": f"{staff.first_name} {staff.last_name}"} for staff in staffs
        ]

        return JsonResponse({"subjects": subject_data, "staffs": staff_data}, status=200)
    
    return JsonResponse({"subjects": [], "staffs": []}, status=400)
    
def save_load_data(request):
    if request.method == 'POST':
        try:
            # Parse the JSON data from the request body
            data = json.loads(request.body)

            # Extract 'loads' list from the data
            loads = data.get('loads', [])
            for load_data in loads:
                # Fetch the related model instances based on the provided IDs
                session_year = SessionYearModel.objects.get(id=load_data['session_year_id'])
                curriculum = Curriculums.objects.get(id=load_data['curriculum_id'])
                assign_section = AssignSection.objects.get(id=load_data['assign_section_id'])
                subject = Subjects.objects.get(id=load_data['subject_id'])
                staff = CustomUser.objects.get(id=load_data['staff_id'])

                # Create the Load record
                Load.objects.create(
                    session_year_id=session_year,
                    curriculum_id=curriculum,
                    AssignSection_id=assign_section,
                    subject_id=subject,
                    staff_id=staff,
                    is_advisory=load_data.get('is_advisory', False)  # Default to False if not provided
                )

            # Set a success message
            messages.success(request, "Load data saved successfully!")

            # Return a JSON response with success message and redirect URL
            return JsonResponse({
                "message": "Load data saved successfully!",
                "redirect_url": "/add_load/"  # Update to your desired redirect path
            })
        except Exception as e:
            print("Error:", e)  # Log any error for debugging
            return JsonResponse({"message": str(e)}, status=400)

    return JsonResponse({"message": "Invalid request method"}, status=405)

def add_load_save(request):
    if request.method != "POST":
        messages.error(request, "Method Not Allowed!")
        return redirect('add_load')
    else:
        try:
            # Retrieve form data
            session_year_id = request.POST.get('session_year_id')
            session_id = SessionYearModel.objects.get(id=session_year_id)

            curriculum_id = request.POST.get('curriculum_id')
            curriculum = Curriculums.objects.get(id=curriculum_id)

            assignsection_id = request.POST.get('AssignSection_id')  # Adjusted to use AssignSection_id
            assignsection = AssignSection.objects.get(id=assignsection_id)

            subject_id = request.POST.get('subject_id')
            subject = Subjects.objects.get(id=subject_id)

            staff_id = request.POST.get('staff_id')
            staff_user = CustomUser.objects.get(id=staff_id)

            is_advisory = request.POST.get('is_advisory')

            # Check if the subject is already assigned to this section in the same session
            conflicting_load = Load.objects.filter(
                session_year_id=session_id,
                AssignSection_id=assignsection,
                subject_id=subject
            ).first()  # Get the first conflicting load record

            if conflicting_load:
                conflict_details = (
                    f"Conflict detected: {assignsection.GradeLevel_id.GradeLevel_name} - {assignsection.section_id.section_name} "
                    f"already has {conflicting_load.staff_id.first_name} "
                    f"{conflicting_load.staff_id.last_name} assigned for subject {subject.subject_name}."
                )
                messages.warning(request, f"WARNING!!! {conflict_details}")
                return redirect('add_load')

            # Check for duplicate load entry with the same staff
            if Load.objects.filter(
                session_year_id=session_id,
                curriculum_id=curriculum,
                AssignSection_id=assignsection,
                subject_id=subject,
                staff_id=staff_user
            ).exists():
                messages.warning(request, "WARNING!!! This load record already exists for this faculty.")
                return redirect('add_load')

            # Get the related Staff instance
            staff = Staffs.objects.get(admin=staff_user)

            # Count the current number of loads for the selected academic year
            current_load = Load.objects.filter(
                session_year_id=session_id,
                staff_id=staff_user
            ).count()

            # Check if the current load exceeds or equals the staff's max_load for the selected academic year
            if current_load >= staff.max_load:
                messages.error(request, f"{staff_user.first_name} {staff_user.last_name} has reached the maximum load limit of {staff.max_load} for the academic year {session_id.session_start_year} - {session_id.session_end_year}.")
                return redirect('add_load')

            # Saving the load
            load = Load(
                session_year_id=session_id,
                curriculum_id=curriculum,
                AssignSection_id=assignsection,
                subject_id=subject,
                staff_id=staff_user,
                is_advisory=is_advisory
            )
            load.save()

            messages.success(request, "Load Added Successfully!")
            return redirect('add_load')

        except Exception as e:
            messages.error(request, f"Failed to Add Load! Error: {e}")
            return redirect('add_load')



def add_schedule(request):
    gradelevels = GradeLevel.objects.all()
    sections = Section.objects.all()
    loads = Load.objects.all()  
    context = {
        'gradelevels': gradelevels,
        'sections': sections,
        "loads": loads,  
    }
    return render(request, 'hod_template/Add_Template/add_schedule_template.html', context)


def add_schedule_save(request):
    if request.method != "POST":
        messages.error(request, "Invalid Method")
        return redirect('add_schedule')
    else:
        try:
            # Retrieve form data
            session_year_id = request.POST.get('session_year_id')
            session_id = SessionYearModel.objects.get(id=session_year_id)
            
            staff_id = request.POST.get('staff_id')
            staff = CustomUser.objects.get(id=staff_id)

            load_id = request.POST.get('load_id')
            load = Load.objects.get(id=load_id)

            day_of_week = request.POST.get('day_of_week')
            start_time = request.POST.get('start_time')
            end_time = request.POST.get('end_time')

            # Convert start and end times to Python time objects
            start_time_obj = time.fromisoformat(start_time)
            end_time_obj = time.fromisoformat(end_time)

            # Set the allowed time range
            allowed_start = time(7, 0)  # 7:00 AM
            allowed_end = time(17, 0)   # 5:00 PM

            # Check if the times fall within the allowed range
            if not (allowed_start <= start_time_obj < allowed_end) or not (allowed_start < end_time_obj <= allowed_end):
                messages.warning(request, "Schedule times must be between 8:00 AM and 3:00 PM.")
                return redirect('add_schedule')

            # Check for any overlapping schedules
            conflicting_schedules = Schedule.objects.filter(
                session_year_id=session_id,
                staff_id=staff,
                day_of_week=day_of_week
            ).filter(
                Q(start_time__lt=end_time_obj) & Q(end_time__gt=start_time_obj)
            )

            if conflicting_schedules.exists():
                # Get details of the first conflicting schedule for the message
                conflict = conflicting_schedules.first()
                conflict_details = (
                    f"Staff {staff.first_name} {staff.last_name} already has a schedule for "
                    f"{conflict.load_id.subject_id.subject_name} on {day_of_week} from "
                    f"{conflict.start_time.strftime('%I:%M %p')} to {conflict.end_time.strftime('%I:%M %p')}."
                )

                messages.warning(request, f"WARNING!!! Schedule conflict detected: {conflict_details}")
                return redirect('add_schedule')

            # Save the new schedule if no conflicts
            schedule = Schedule(
                session_year_id=session_id,
                staff_id=staff, 
                load_id=load,
                day_of_week=day_of_week,
                start_time=start_time,
                end_time=end_time
            )
            schedule.save()

            messages.success(request, "Schedule Added Successfully!")
            return redirect('add_schedule')
        
        except Exception as e:
            print(e)  # For debugging
            messages.error(request, f"Failed to Add Schedule! Error: {e}")
            return redirect('add_schedule')

    # Return the data as a JSON response
    return JsonResponse(load_data, safe=False)

def manage_load_scheduling(request):
    loads = Load.objects.all()
    context = {
        "loads": loads,
    }
    return render(request, 'hod_template/Manage_Template/manage_load_template.html', context)

def manage_class_scheduling(request):
    gradelevels = GradeLevel.objects.all()
    sections = Section.objects.all()
    schedules = Schedule.objects.all()
    context = {
        "gradelevels": gradelevels,
        "sections": sections,
        "schedules": schedules,
    }
    return render(request, 'hod_template/Manage_Template/manage_schedule_template.html', context)

# FETCH LOAD DATA
def fetch_load_data(request):                                   
    grade_level_id = request.GET.get('gradelevel_id')
    section_id = request.GET.get('section_id')

    if grade_level_id and section_id:
        # Get the load_ids that are already scheduled
        scheduled_load_ids = Schedule.objects.values_list('load_id', flat=True)

        # Filter out the loads that are already scheduled
        loads = Load.objects.filter(
            AssignSection_id__GradeLevel_id=grade_level_id,
            AssignSection_id__section_id=section_id
        ).exclude(id__in=scheduled_load_ids).select_related('AssignSection_id', 'subject_id', 'staff_id')
        
        load_data = [
            {
                'id': load.id,
                'session_year': f"{load.session_year_id.session_start_year.year}-{load.session_year_id.session_end_year.year}",
                'subject': load.subject_id.subject_name,
                'staff': f"{load.staff_id.first_name} {load.staff_id.last_name}",
                'grade_level': load.AssignSection_id.GradeLevel_id.GradeLevel_name,
                'section': load.AssignSection_id.section_id.section_name,
            }
            for load in loads
        ]
        
        return JsonResponse({'loads': load_data}, safe=False)

    return JsonResponse({'loads': []}, safe=False)

def check_schedule_conflict(request):
    if request.method == 'GET':
        day_of_week = request.GET.get('day_of_week')
        start_time = request.GET.get('start_time')
        end_time = request.GET.get('end_time')
        gradelevel_id = request.GET.get('gradelevel_id')
        section_id = request.GET.get('section_id')
        session_year_id = request.GET.get('session_year_id')  # Added session_year_id

        # Ensure all required data is provided
        if not (day_of_week and start_time and end_time and gradelevel_id and section_id and session_year_id):
            return JsonResponse({"conflict": False, "message": "Incomplete data provided."})

        # Convert times to time objects
        start_time = time.fromisoformat(start_time)
        end_time = time.fromisoformat(end_time)

        # Check for schedule conflicts in the database for the given grade level, section, and session year
        conflicts = Schedule.objects.filter(
            load_id__AssignSection_id__GradeLevel_id=gradelevel_id,  # Check for grade level conflict
            load_id__AssignSection_id__section_id=section_id,       # Check for section conflict
            load_id__session_year_id=session_year_id,                # Check for session year conflict
            day_of_week=day_of_week,
        ).filter(
            Q(start_time__lt=end_time) & Q(end_time__gt=start_time)  # Overlapping condition
        )

        # If conflicts exist, return conflict details
        if conflicts.exists():
            # Get the first conflicting schedule for detailed response
            conflicting_schedule = conflicts.first()
            
            # Format the start and end times to include AM/PM
            formatted_start_time = conflicting_schedule.start_time.strftime("%I:%M %p")  # Example: 02:30 PM
            formatted_end_time = conflicting_schedule.end_time.strftime("%I:%M %p")      # Example: 03:20 PM
            
            return JsonResponse({
                "conflict": True,
                "message": (
                    f"Time conflict detected with existing schedule for "
                    f"{conflicting_schedule.load_id.subject_id.subject_name} taught by {conflicting_schedule.load_id.staff_id.first_name} "
                    f"on {conflicting_schedule.day_of_week} from {formatted_start_time} to {formatted_end_time}."
                )
            })

        # No conflicts found
        return JsonResponse({"conflict": False, "message": "No conflicts detected."})

    # Return a 405 Method Not Allowed for non-GET requests
    return JsonResponse({"error": "Invalid request method. Only GET is allowed."}, status=405)

@csrf_exempt
def save_schedule(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            schedules = data.get('schedules', [])
            for schedule in schedules:
                load = Load.objects.get(id=schedule['load_id'])

                # Iterate through each day in the list of selected days
                for day in schedule['day_of_week']:
                    # Create a schedule entry for each day
                    Schedule.objects.create(
                        load_id=load,
                        day_of_week=day,  # Use each day of the week separately
                        start_time=schedule['start_time'],
                        end_time=schedule['end_time']
                    )

            messages.success(request, "Schedules saved successfully!")
            return JsonResponse({
                "message": "Schedules saved successfully!",
                "redirect_url": "/add_schedule/"
            })

        except Exception as e:
            print(e)
            return JsonResponse({"message": str(e)}, status=400)

    return JsonResponse({"message": "Invalid request method"}, status=405)

def search_schedule(request):
    query = request.GET.get('query', '').strip()  # Get the query and trim spaces

    # Redirect to the schedule management page if the query is empty
    if not query:
        return redirect('manage_class_scheduling')  # Use the name of your URL pattern for manage_class_scheduling

    # Try to parse the time query (e.g., '08:00 a.m.')
    time_query = None
    try:
        # Convert query to 24-hour time format
        time_query = dt.strptime(query, '%I:%M %p').time()  # '%I:%M %p' is the format for '08:00 a.m.'
    except ValueError:
        pass  # If the time format doesn't match, just ignore

    if time_query:
        # If time query is valid, filter by start_time and end_time
        schedules = Schedule.objects.filter(
            Q(start_time__exact=time_query) | Q(end_time__exact=time_query)
        )
    else:
        # Otherwise, perform search based on other criteria
        schedules = Schedule.objects.filter(
            Q(load_id__subject_id__subject_name__icontains=query) |
            Q(load_id__staff_id__first_name__icontains=query) |
            Q(load_id__staff_id__last_name__icontains=query) |
            Q(load_id__AssignSection_id__GradeLevel_id__GradeLevel_name__icontains=query) |
            Q(load_id__AssignSection_id__section_id__section_name__icontains=query)
        )

    gradelevels = GradeLevel.objects.all()
    sections = Section.objects.all()

    context = {
        "gradelevels": gradelevels,
        "sections": sections,
        "schedules": schedules,
        "search_query": query,  # Pass the search query to the template
    }

    return render(request, 'hod_template/Manage_Template/manage_schedule_template.html', context)
    
def edit_schedule(request, schedule_id):
    # Adding Schedule ID into Session Variable
    request.session['schedule_id'] = schedule_id

    try:
        # Retrieve the schedule object using the provided schedule_id
        schedule = Schedule.objects.get(id=schedule_id)
    except Schedule.DoesNotExist:
        messages.error(request, "Schedule does not exist.")
        return redirect('manage_class_scheduling')  # Adjust this redirect based on your URL names

    # Initialize the form with data from the database
    form = EditScheduleForm()
    form.fields['GradeLevel_id'].initial = schedule.GradeLevel_id.id
    form.fields['subject_id'].initial = schedule.subject_id.id
    form.fields['staff_id'].initial = schedule.staff_id.admin.id
    form.fields['session_year_id'].initial = schedule.session_year_id.id
    form.fields['day_of_week'].initial = schedule.day_of_week
    form.fields['start_time'].initial = schedule.start_time
    form.fields['end_time'].initial = schedule.end_time

    context = {
        "id": schedule_id,
        "form": form
    }

    return render(request, "hod_template/Edit_Template/edit_schedule_template.html", context)

def edit_schedule_save(request):
    if request.method != "POST":
        return HttpResponse("Invalid Method!")
    else:
        schedule_id = request.session.get('schedule_id')
        if schedule_id is None:
            return redirect('/manage_class_scheduling')

        form = EditScheduleForm(request.POST, request.FILES)
        if form.is_valid():
            subject_id = form.cleaned_data['subject_id']
            GradeLevel_id = form.cleaned_data['GradeLevel_id']
            staff_id = form.cleaned_data['staff_id']
            session_year_id = form.cleaned_data['session_year_id']
            day_of_week = form.cleaned_data['day_of_week']
            start_time = form.cleaned_data['start_time']
            end_time = form.cleaned_data['end_time']

            try:
                # Update Schedule model
                schedule = Schedule.objects.get(id=schedule_id)
                
                # Update schedule fields
                schedule.subject_id = Subjects.objects.get(subject_id)
                schedule.GradeLevel_id = GradeLevel.objects.get(id=GradeLevel_id)
                schedule.staff_id = Staffs.objects.get(id=staff_id)
                schedule.session_year_id = SessionYearModel.objects.get(id=session_year_id)
                schedule.day_of_week = day_of_week
                schedule.start_time = start_time
                schedule.end_time = end_time
                schedule.save()
                
                # Clear the session variable
                del request.session['schedule_id']

                messages.success(request, "Schedule Updated Successfully!")
                return redirect(f'/edit_schedule/{schedule_id}')
            except Exception as e:
                messages.error(request, f"Failed to Update Schedule: {e}")
                return redirect(f'/edit_schedule/{schedule_id}')
        else:
            return redirect(f'/edit_schedule/{schedule_id}')




@csrf_exempt
def check_email_exist(request):
    email = request.POST.get("email")
    user_obj = CustomUser.objects.filter(email=email).exists()
    if user_obj:
        return HttpResponse(True)
    else:
        return HttpResponse(False)


@csrf_exempt
def check_username_exist(request):
    username = request.POST.get("username")
    user_obj = CustomUser.objects.filter(username=username).exists()
    if user_obj:
        return HttpResponse(True)
    else:
        return HttpResponse(False)



def student_feedback_message(request):
    feedbacks = FeedBackStudent.objects.all()
    context = {
        "feedbacks": feedbacks
    }
    return render(request, 'hod_template/student_feedback_template.html', context)


@csrf_exempt
def student_feedback_message_reply(request):
    feedback_id = request.POST.get('id')
    feedback_reply = request.POST.get('reply')

    try:
        feedback = FeedBackStudent.objects.get(id=feedback_id)
        feedback.feedback_reply = feedback_reply
        feedback.save()
        return HttpResponse("True")

    except:
        return HttpResponse("False")


def staff_feedback_message(request):
    feedbacks = FeedBackStaffs.objects.all()
    context = {
        "feedbacks": feedbacks
    }
    return render(request, 'hod_template/staff_feedback_template.html', context)


@csrf_exempt
def staff_feedback_message_reply(request):
    feedback_id = request.POST.get('id')
    feedback_reply = request.POST.get('reply')

    try:
        feedback = FeedBackStaffs.objects.get(id=feedback_id)
        feedback.feedback_reply = feedback_reply
        feedback.save()
        return HttpResponse("True")

    except:
        return HttpResponse("False")


def student_leave_view(request):
    leaves = LeaveReportStudent.objects.all()
    context = {
        "leaves": leaves
    }
    return render(request, 'hod_template/student_leave_view.html', context)

def student_leave_approve(request, leave_id):
    leave = LeaveReportStudent.objects.get(id=leave_id)
    leave.leave_status = 1
    leave.save()
    return redirect('student_leave_view')


def student_leave_reject(request, leave_id):
    leave = LeaveReportStudent.objects.get(id=leave_id)
    leave.leave_status = 2
    leave.save()
    return redirect('student_leave_view')


def staff_leave_view(request):
    leaves = LeaveReportStaff.objects.all()
    context = {
        "leaves": leaves
    }
    return render(request, 'hod_template/staff_leave_view.html', context)


def staff_leave_approve(request, leave_id):
    leave = LeaveReportStaff.objects.get(id=leave_id)
    leave.leave_status = 1
    leave.save()
    return redirect('staff_leave_view')


def staff_leave_reject(request, leave_id):
    leave = LeaveReportStaff.objects.get(id=leave_id)
    leave.leave_status = 2
    leave.save()
    return redirect('staff_leave_view')


def admin_view_attendance(request):
    subjects = Subjects.objects.all()
    session_years = SessionYearModel.objects.all()
    context = {
        "subjects": subjects,
        "session_years": session_years
    }
    return render(request, "hod_template/admin_view_attendance.html", context)


@csrf_exempt
def admin_get_attendance_dates(request):
    # Getting Values from Ajax POST 'Fetch Student'
    subject_id = request.POST.get("subject")
    session_year = request.POST.get("session_year_id")

    # Students enroll to GradeLevel, GradeLevel has Subjects
    # Getting all data from subject model based on subject_id
    subject_model = Subjects.objects.get(id=subject_id)

    session_model = SessionYearModel.objects.get(id=session_year)

    # students = Students.objects.filter(GradeLevel_id=subject_model.GradeLevel_id, session_year_id=session_model)
    attendance = Attendance.objects.filter(subject_id=subject_model, session_year_id=session_model)

    # Only Passing Student Id and Student Name Only
    list_data = []

    for attendance_single in attendance:
        data_small={"id":attendance_single.id, "attendance_date":str(attendance_single.attendance_date), "session_year_id":attendance_single.session_year_id.id}
        list_data.append(data_small)

    return JsonResponse(json.dumps(list_data), content_type="application/json", safe=False)


@csrf_exempt
def admin_get_attendance_student(request):
    # Getting Values from Ajax POST 'Fetch Student'
    attendance_date = request.POST.get('attendance_date')
    attendance = Attendance.objects.get(id=attendance_date)

    attendance_data = AttendanceReport.objects.filter(attendance_id=attendance)
    # Only Passing Student Id and Student Name Only
    list_data = []

    for student in attendance_data:
        data_small={"id":student.student_id.admin.id, "name":student.student_id.admin.first_name+" "+student.student_id.admin.last_name, "status":student.status}
        list_data.append(data_small)

    return JsonResponse(json.dumps(list_data), content_type="application/json", safe=False)


def admin_profile(request):
    user = CustomUser.objects.get(id=request.user.id)

    context={
        "user": user
    }
    return render(request, 'hod_template/admin_profile.html', context)


def admin_profile_update(request):
    if request.method != "POST":
        messages.error(request, "Invalid Method!")
        return redirect('admin_profile')
    else:
        first_name = request.POST.get('first_name')
        last_name = request.POST.get('last_name')
        password = request.POST.get('password')

        try:
            customuser = CustomUser.objects.get(id=request.user.id)
            customuser.first_name = first_name
            customuser.last_name = last_name
            if password != None and password != "":
                customuser.set_password(password)
            customuser.save()
            messages.success(request, "Profile Updated Successfully")
            return redirect('admin_profile')
        except:
            messages.error(request, "Failed to Update Profile")
            return redirect('admin_profile')
    


def staff_profile(request):
    pass


def student_profile(requtest):
    pass



